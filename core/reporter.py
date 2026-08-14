"""
SAP Migration Post-Load Validator — Excel Report Generator V4
Accepts a ValidationResult dataclass or the result_dict from app.py.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path


def _coerce(result) -> dict:
    if isinstance(result, dict):
        return result
    ss      = result.summary_stats
    mapping = result.mapping
    thr     = mapping.pass_threshold if mapping else 100.0
    sel     = mapping.selected_fields if mapping else []
    rows    = []
    for fr in result.field_results:
        rows.append({
            "field":          fr.field_source,
            "field_label":    fr.field_label,
            "field_target":   fr.field_target,
            "type":           "numeric" if fr.is_numeric else "string",
            "tolerance":      fr.tolerance_used,
            "total":          fr.total_records,
            "matched":        fr.matched,
            "mismatched":     fr.mismatched,
            "miss_source":    fr.missing_in_source,
            "miss_target":    fr.missing_in_target,
            "match_pct":      fr.match_pct,
            "pass_threshold": fr.pass_threshold,
            "status":         fr.status,
            "mismatches":     fr.mismatch_details,
            "matches":        fr.matched_details,
            "display_name":   fr.field_label,
            "is_cross_mapped": fr.field_source != fr.field_target,
        })
    mp = None
    if mapping:
        mp = {
            "join_key":           mapping.join_key,
            "join_key_label":     mapping.join_key_label,
            "numeric_fields":     mapping.numeric_fields,
            "tolerance_map":      mapping.tolerance_map,
            "source_only_fields": mapping.source_only_fields,
            "target_only_fields": mapping.target_only_fields,
            "selected_fields":    sel,
            "pass_threshold":     thr,
        }
    return {
        "name":                   Path(result.source_file).stem.upper(),
        "status":                 result.overall_status,
        "source_file":            result.source_file,
        "target_file":            result.target_file,
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "errors":                 result.errors,
        "mapping":                mp,
        "field_results":          rows,
        "run_at":                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def generate_excel_report(result, output_path: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("pip install openpyxl")

    r        = _coerce(result)
    mapping  = r.get("mapping") or {}
    pass_thr = mapping.get("pass_threshold", 100.0)
    sel      = mapping.get("selected_fields", [])

    C_NAVY  = "FF1B3A57"; C_WHITE  = "FFFFFFFF"
    C_GREEN = "FF00AA44"; C_RED    = "FFCC2200"
    C_AMBER = "FFDD8800"; C_DARK   = "FF333333"
    C_LG    = "FFE6F4EA"; C_LR     = "FFFCE8E6"; C_LGREY = "FFF5F5F5"

    def fill(c):  return PatternFill("solid", fgColor=c)
    def bdr():
        s = Side(style="thin", color="FFCCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def hcell(ws, row, col, val, bg=C_NAVY):
        c = ws.cell(row, col, val)
        c.fill = fill(bg); c.font = Font(bold=True, color=C_WHITE, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr(); return c

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A23"

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"SAP Post-Load Validation — {r['name']}"
    c.font  = Font(bold=True, size=16, color=C_WHITE)
    c.fill  = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    oc = C_GREEN if r["status"] == "PASS" else C_RED
    meta = [
        ("Run Date",       r.get("run_at", "")),
        ("Source File",    r["source_file"]),
        ("Target File",    r["target_file"]),
        ("Pass Threshold", f">= {pass_thr}%"),
        ("Fields Scope",   f"{len(sel)} selected" if sel else "All common fields"),
        ("Overall Status", r["status"]),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(i, 1, k).font = Font(bold=True, color="FF444444")
        cell = ws.cell(i, 2, v)
        if k == "Overall Status": cell.font = Font(bold=True, color=oc, size=12)
        if k == "Pass Threshold": cell.font = Font(bold=True, color=C_AMBER)

    ws.cell(11, 1, "Record Counts").font = Font(bold=True, size=11)
    for i, (k, v) in enumerate([
        ("Source Records", r["total_source_records"]),
        ("Target Records", r["total_target_records"]),
        ("Keys Matched",   r["records_matched"]),
        ("Source Only",    r["records_only_in_source"]),
        ("Target Only",    r["records_only_in_target"]),
    ], start=12):
        ws.cell(i, 1, k)
        cell = ws.cell(i, 2, v)
        if k in ("Source Only", "Target Only") and v > 0:
            cell.font = Font(bold=True, color=C_RED)

    ws.cell(11, 4, "Validation Stats").font = Font(bold=True, size=11)
    for i, (k, v) in enumerate([
        ("Fields Validated", r["total_fields"]),
        ("Fields Passed",    r["fields_passed"]),
        ("Fields Failed",    r["fields_failed"]),
        ("Pass Rate",        f"{r['pass_rate_pct']}%"),
    ], start=12):
        ws.cell(i, 4, k)
        cell = ws.cell(i, 5, v)
        if k == "Fields Failed" and isinstance(v, int) and v > 0:
            cell.font = Font(bold=True, color=C_RED)
        if k == "Pass Rate":
            cell.font = Font(bold=True, color=C_GREEN if r["fields_failed"] == 0 else C_RED)

    if mapping:
        ws.cell(11, 7, "Auto-Detected").font = Font(bold=True, size=11)
        ws.cell(12, 7, "Join Key")
        ws.cell(12, 8, f"{mapping.get('join_key_label', '')}  ({mapping.get('join_key', '')})")
        ws.cell(13, 7, "Numeric Fields")
        ws.cell(13, 8, ", ".join(mapping.get("numeric_fields", [])) or "none")

    ws.cell(19, 1, "Field-Level Results").font = Font(bold=True, size=11)
    ws.cell(20, 1, f"Pass threshold: >= {pass_thr}%").font = Font(color=C_AMBER, size=10, italic=True)

    hdrs = ["Field Label","Source Field","Target Field","Map Method","Type","Tolerance",
            "Total","Matched","Mismatched","Miss-Src","Miss-Tgt","Match %","Threshold","Status"]
    for col, h in enumerate(hdrs, 1):
        hcell(ws, 22, col, h)

    for ri, fr in enumerate(r["field_results"], start=23):
        thr    = fr.get("pass_threshold", pass_thr)
        pct    = fr["match_pct"]
        status = fr["status"]
        bg     = C_LG if status == "PASS" else C_LR
        tol    = f"+-{fr['tolerance']}" if fr.get("tolerance") is not None else "—"
        label  = fr.get("display_name") or fr.get("field_label") or fr["field"]
        method = fr.get("mapping_method", "exact")
        vals   = [label, fr["field"], fr.get("field_target", fr["field"]), method,
                  fr.get("type",""), tol, fr["total"], fr["matched"], fr["mismatched"],
                  fr["miss_source"], fr["miss_target"], f"{pct}%", f">= {thr}%", status]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(ri, col, val)
            cell.fill = fill(bg); cell.border = bdr()
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
            if col == len(vals): cell.font = Font(bold=True, color=C_GREEN if status == "PASS" else C_RED)
            if col == 12: cell.font = Font(bold=True, color=C_GREEN if pct >= thr else C_RED)
    if r["field_results"]:
        ws.auto_filter.ref = f"A22:N{22 + len(r['field_results'])}"

    col_widths = [26, 14, 14, 14, 8, 10, 8, 10, 12, 10, 10, 10, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for fr in r["field_results"]:
        if fr["status"] != "FAIL" or not fr.get("mismatches"):
            continue
        label  = fr.get("display_name") or fr.get("field_label") or fr["field"]
        safe   = label[:22].replace("/", "_").replace("\\", "_")
        ws2    = wb.create_sheet(title=f"FAIL_{safe}")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:G1")
        c2 = ws2["A1"]
        src_tgt = fr["field"] if fr["field"] == fr.get("field_target", fr["field"]) \
                  else f"{fr['field']} \u2192 {fr.get('field_target', '')}"
        c2.value = f"Mismatches — {label}  ({src_tgt})"
        c2.font  = Font(bold=True, size=13, color=C_WHITE)
        c2.fill  = fill(C_RED)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28

        ws2.merge_cells("A2:G2")
        ws2["A2"].value = (f"Match: {fr['match_pct']}%  |  Threshold: >= {thr}%  |  "
                           f"Total: {fr['total']}  |  Matched: {fr['matched']}")
        ws2["A2"].font  = Font(size=10, color="FF555555")
        ws2["A2"].fill  = fill(C_LR)
        ws2["A2"].alignment = Alignment(horizontal="center")

        field_rec = next((item for item in r.get("recommendations", [])
                          if item.get("field") == fr.get("field") or
                          item.get("target_field") == fr.get("field_target")), None)

        for col, h in enumerate(["Key", "Source Value", "Target Value", "Issue",
                                 "Source Field", "Target Field", "Recommendation"], 1):
            hcell(ws2, 4, col, h, bg=C_DARK)

        for ri2, mismatch in enumerate(fr["mismatches"], start=5):
            bg2 = C_LR if ri2 % 2 == 0 else C_LGREY
            issue = str(mismatch.get("issue", ""))
            source_value = mismatch.get("source_value", "")
            target_value = mismatch.get("target_value", "")
            if issue == "Missing in source":
                row_recommendation = ("Review the source extraction. A target value exists, but the source is blank; "
                                      "do not replace automatically.")
            elif issue == "Missing in target":
                row_recommendation = f"Fill the target with the validated source value: {source_value}"
            elif issue.startswith("Numeric delta"):
                row_recommendation = (f"Replace target value {target_value} with source value {source_value} "
                                      "after confirming the configured tolerance.")
            else:
                row_recommendation = f"Replace target value '{target_value}' with source value '{source_value}'."
            if field_rec and field_rec.get("rag_match"):
                row_recommendation += (f" Learned rule match: {field_rec.get('rag_confidence', 0)}% "
                                       f"({field_rec.get('approved_count', 0)} prior approval(s)).")
            for ci, v in enumerate([
                mismatch.get("material",""), source_value, target_value, issue,
                fr["field"], fr.get("field_target", fr["field"]), row_recommendation
            ], 1):
                cell = ws2.cell(ri2, ci, v)
                cell.fill = fill(bg2); cell.border = bdr()
                cell.alignment = Alignment(vertical="top", wrap_text=(ci == 7))
                if ci == 2: cell.font = Font(color=C_GREEN)
                if ci == 3: cell.font = Font(color=C_RED)

        for col, w in zip("ABCDEFG", [24, 30, 30, 30, 14, 14, 58]):
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = "A5"
        ws2.auto_filter.ref = f"A4:G{4 + len(fr['mismatches'])}"

        # Keep the recommendation with the failed field instead of creating a
        # separate workbook sheet that forces users to cross-reference fields.
        rec = field_rec
        if rec:
            ws2.merge_cells("H1:L1")
            ws2["H1"] = "Recommended Correction"
            ws2["H1"].font = Font(bold=True, size=12, color=C_WHITE)
            ws2["H1"].fill = fill(C_AMBER)
            ws2["H1"].alignment = Alignment(horizontal="center")
            details = [
                ("Field", rec.get("label") or fr.get("field_label") or fr.get("field")),
                ("Target field", rec.get("target_field") or fr.get("field_target")),
                ("Severity", str(rec.get("severity", "")).upper()),
                ("Affected records", rec.get("affected_records", 0)),
                ("Current match", f"{rec.get('match_pct', 0)}%"),
                ("Recommendation", rec.get("explanation", "Review mismatch details.")),
                ("Dashboard action", "Create corrected target copy" if rec.get("can_apply") else "Manual review required"),
                ("Learned rule", "YES" if rec.get("learned") else "NO"),
                ("RAG confidence", f"{rec.get('rag_confidence', 0)}%" if rec.get("rag_match") else "Not matched"),
            ]
            for row, (key, value) in enumerate(details, 3):
                label_cell = ws2.cell(row, 8, key)
                value_cell = ws2.cell(row, 9, value)
                label_cell.font = Font(bold=True, color=C_DARK)
                label_cell.fill = fill("FFFFF4D6")
                value_cell.fill = fill("FFFFFBEB")
                label_cell.border = value_cell.border = bdr()
                value_cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws2.column_dimensions["H"].width = 20
            ws2.column_dimensions["I"].width = 48

    # Consolidate all successful fields into one matched-record worksheet.
    passed_rows = []
    for fr in r["field_results"]:
        matches = fr.get("matches", [])
        if fr.get("status") != "PASS" or not matches:
            continue
        label = fr.get("display_name") or fr.get("field_label") or fr["field"]
        for match in matches:
            passed_rows.append([
                label, fr["field"], fr.get("field_target", fr["field"]),
                match.get("material", ""), match.get("source_value", ""),
                match.get("target_value", ""), match.get("result", "Exact match"),
            ])

    if passed_rows:
        wsp = wb.create_sheet(title="Passed Records")
        wsp.sheet_view.showGridLines = False
        wsp.merge_cells("A1:G1")
        wsp["A1"] = "Consolidated Passed Records"
        wsp["A1"].font = Font(bold=True, size=13, color=C_WHITE)
        wsp["A1"].fill = fill(C_GREEN)
        wsp["A1"].alignment = Alignment(horizontal="center")
        wsp.merge_cells("A2:G2")
        wsp["A2"] = (f"All {len(passed_rows):,} matched field-record results across "
                      f"{sum(1 for fr in r['field_results'] if fr.get('status') == 'PASS')} passed fields")
        wsp["A2"].fill = fill(C_LG)
        wsp["A2"].alignment = Alignment(horizontal="center")
        headers = ["Field", "Source Field", "Target Field", "Key", "Source Value", "Target Value", "Result"]
        for col, header in enumerate(headers, 1):
            hcell(wsp, 4, col, header, bg=C_DARK)
        for row, values in enumerate(passed_rows, 5):
            for col, value in enumerate(values, 1):
                cell = wsp.cell(row, col, value)
                cell.fill = fill(C_LG if row % 2 == 0 else C_LGREY)
                cell.border = bdr()
                if col in (5, 6): cell.font = Font(color=C_GREEN)
        for col, width in zip("ABCDEFG", [26, 16, 16, 24, 30, 30, 22]):
            wsp.column_dimensions[col].width = width
        wsp.freeze_panes = "A5"
        wsp.auto_filter.ref = f"A4:G{4 + len(passed_rows)}"

    wb.save(output_path)
    return output_path
