"""
Genpact SAP Migration Validator — Dashboard V4
Run:  python dashboard/app.py
Open: http://localhost:5000
"""

import sys
import os
import pandas as pd
import csv as csv_mod
import threading
import time
import json
import io
import html as html_lib
from urllib.parse import quote
from pathlib import Path
from datetime import datetime

from flask import Flask, render_template, jsonify, send_file, request, Response
from werkzeug.utils import secure_filename

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.validator    import MaterialValidator
from core.reporter     import generate_excel_report
from core.field_labels import (
    get_label, load_custom_labels, enrich_field_rows,
    get_display, SAP_FIELD_LABELS
)
from core.field_mapper  import build_field_mapping, mapping_result_to_dict
from core.object_config import get_object_config, SAP_OBJECT_CONFIG
from core.correction_rag import CorrectionRAG

app = Flask(__name__)

BASE_DIR      = Path(__file__).parent.parent
REPORTS_DIR   = BASE_DIR / "reports"
CONFIG_FILE   = BASE_DIR / "config.json"
LABELS_FILE   = BASE_DIR / "custom_labels.csv"
MAPPING_FILE  = BASE_DIR / "custom_mapping.csv"   # user-defined source->target pairs
TEMPLATES_DIR = BASE_DIR / "templates"
LEARNED_RULES_FILE = BASE_DIR / "config" / "learned_corrections.json"
CORRECTION_MEMORY_FILE = BASE_DIR / "config" / "correction_memory.json"

REPORTS_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_CONFIG = {
    "source_dir":      str(BASE_DIR / "data" / "source"),
    "target_dir":      str(BASE_DIR / "data" / "target"),
    "pass_threshold":  100.0,
    "selected_fields": [],
    "manual_pairs":    [],
    "active_template": "",
}

results_store = {}
row_details_store = {}
scan_status   = {
    "last_scan": None, "scanning": False, "error": None,
    "current_file": None, "total_files": 0, "completed_files": 0,
}
file_states  = {}
activity_log = []

SUPPORTED_EXT = {".csv", ".xlsx", ".xls"}
TEMPLATE_EXT  = {".csv", ".xlsx", ".xls", ".txt"}
scan_lock     = threading.Lock()
correction_rag = CorrectionRAG(CORRECTION_MEMORY_FILE)


def load_learned_rules():
    """Return user-approved correction rules. Rules are suggestions, never auto-applied."""
    try:
        return json.loads(LEARNED_RULES_FILE.read_text()) if LEARNED_RULES_FILE.exists() else []
    except Exception:
        return []


def save_learned_rule(object_name, source_field, target_field):
    rules = load_learned_rules()
    key = (object_name.upper(), source_field.upper(), target_field.upper())
    rules = [r for r in rules if (
        r.get("object", "").upper(), r.get("source_field", "").upper(),
        r.get("target_field", "").upper()) != key]
    rules.append({
        "object": key[0], "source_field": key[1], "target_field": key[2],
        "action": "copy_source_to_target", "approved_at": datetime.now().isoformat(timespec="seconds"),
    })
    LEARNED_RULES_FILE.parent.mkdir(parents=True, exist_ok=True)
    LEARNED_RULES_FILE.write_text(json.dumps(rules, indent=2))


def build_recommendations(result):
    """Build conservative, evidence-backed actions for failed non-key columns."""
    learned = load_learned_rules()
    recommendations = []
    for fr in result.get("field_results", []):
        if fr.get("status") != "FAIL" or fr.get("is_key_field"):
            continue
        source_field = fr.get("field", "")
        target_field = fr.get("field_target") or source_field
        samples = fr.get("mismatches", [])
        issues = {m.get("issue", "") for m in samples}
        if issues == {"Missing in source"}:
            action = "Review source extraction; target has values but source is blank."
            can_apply = False
        else:
            action = "Copy validated source values into the matching target records."
            can_apply = True
        was_learned = any(
            r.get("object") == result.get("name") and
            r.get("source_field") == source_field and r.get("target_field") == target_field
            for r in learned
        )
        retrieved = correction_rag.retrieve(
            object_name=result.get("name", ""), source_field=source_field,
            target_field=target_field, issues=list(issues), examples=samples,
        )
        if was_learned and not retrieved:
            correction_rag.remember(
                object_name=result.get("name", ""), source_field=source_field,
                target_field=target_field, issues=list(issues), examples=samples,
            )
            retrieved = correction_rag.retrieve(
                object_name=result.get("name", ""), source_field=source_field,
                target_field=target_field, issues=list(issues), examples=samples,
            )
        best = retrieved[0] if retrieved else None
        retrieval_score = best.get("retrieval_score", 0) if best else 0
        approved_count = best.get("approved_count", 0) if best else 0
        rag_approved = bool(best and best.get("action") == "copy_source_to_target" and retrieval_score >= 0.85)
        recommendations.append({
            "field": source_field, "target_field": target_field,
            "label": fr.get("display_name") or fr.get("field_label") or source_field,
            "severity": "high" if fr.get("match_pct", 0) < 80 else "medium",
            "affected_records": fr.get("mismatched", 0) + fr.get("miss_target", 0),
            "match_pct": fr.get("match_pct", 0), "explanation": action,
            "can_apply": can_apply, "learned": was_learned or rag_approved,
            "rag_match": rag_approved, "rag_confidence": round(retrieval_score * 100, 1),
            "approved_count": approved_count,
            "retrieval_evidence": retrieved,
            "examples": samples[:3],
        })
    return recommendations


def _load_correction_file(path):
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path, dtype=str).fillna("")
    return pd.read_csv(path, dtype=str, encoding="utf-8-sig", keep_default_na=False)


def _normalise_columns(df):
    df = df.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


def find_corrected_files(target_file):
    stem = Path(target_file).stem
    files = list(REPORTS_DIR.glob(f"{stem}_corrected_*")) + list(REPORTS_DIR.glob(f"{stem}_rag_corrected_*"))
    files = sorted(set(files), key=lambda p: p.stat().st_mtime, reverse=True)
    return [{
        "filename": p.name,
        "download_url": "/api/download-corrected/" + p.name,
        "created_at": datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
    } for p in files[:10] if p.suffix.lower() in SUPPORTED_EXT]


def cache_and_trim_row_details(result):
    """Keep complete row details server-side while returning a small dashboard payload."""
    object_name = str(result.get("name", "")).upper()
    cached = {}
    for field_row in result.get("field_results", []):
        field_name = str(field_row.get("field", "")).upper()
        cached[field_name] = {
            "label": field_row.get("display_name") or field_row.get("field_label") or field_name,
            "target_field": field_row.get("field_target") or field_name,
            "matches": field_row.get("matches", []),
            "mismatches": field_row.get("mismatches", []),
        }
        field_row["mismatch_count"] = len(field_row.get("mismatches", []))
        field_row["mismatches"] = field_row.get("mismatches", [])[:20]
        field_row.pop("matches", None)
    row_details_store[object_name] = cached


# ── Config helpers ──────────────────────────────────────────────────────────────

def load_config():
    if CONFIG_FILE.exists():
        try:
            return {**DEFAULT_CONFIG, **json.loads(CONFIG_FILE.read_text())}
        except Exception as e:
            print(f"Config load failed: {e}")
    return dict(DEFAULT_CONFIG)


def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2))


def get_dirs():
    cfg = load_config()
    src = Path(cfg.get("source_dir", DEFAULT_CONFIG["source_dir"]))
    tgt = Path(cfg.get("target_dir", DEFAULT_CONFIG["target_dir"]))
    src.mkdir(parents=True, exist_ok=True)
    tgt.mkdir(parents=True, exist_ok=True)
    return src, tgt


def log_event(message, level="info"):
    entry = {"ts": datetime.now().strftime("%H:%M:%S"), "message": message, "level": level}
    activity_log.append(entry)
    if len(activity_log) > 200:
        activity_log.pop(0)
    print(f"  [{entry['ts']}] [{level.upper()}] {message}")


def cleanup_old_reports(keep=20):
    files = sorted(REPORTS_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    for f in files[keep:]:
        try: f.unlink()
        except: pass


def _get_custom_labels():
    return load_custom_labels(str(LABELS_FILE)) if LABELS_FILE.exists() else {}


def _read_file_headers(src_path: str, tgt_path: str = None):
    """Read only the header row — very fast, no full file load."""
    def headers(path):
        p = Path(path)
        if not p.exists():
            return []
        if p.suffix.lower() in (".xlsx", ".xls"):
            import openpyxl
            wb   = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            ws   = wb.active
            cols = [str(c.value).strip().upper()
                    for c in next(ws.iter_rows(max_row=1)) if c.value]
            wb.close()
            return cols
        with open(str(p), encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            return [c.strip().upper() for c in next(reader)]

    src_cols = headers(src_path) if src_path else []
    tgt_cols = headers(tgt_path) if tgt_path else []
    return src_cols, tgt_cols


# ── Field-selection template helpers ───────────────────────────────────────────

def _read_template_fields(path: Path) -> list:
    """
    Parse a field-selection template file.
    Supported formats:
      CSV  : first column = field names (header row auto-skipped)
      XLSX : column A = field names (header row auto-skipped)
      TXT  : one field name per line (lines starting with # are comments)
    Returns: list of UPPERCASE field names, no blanks, no comments.
    """
    fields = []
    suffix = path.suffix.lower()
    try:
        if suffix == ".txt":
            for line in path.read_text(encoding="utf-8-sig").splitlines():
                val = line.strip()
                if val and not val.startswith("#"):
                    fields.append(val.upper())

        elif suffix in (".xlsx", ".xls"):
            import openpyxl
            wb        = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws        = wb.active
            first_row = True
            for row in ws.iter_rows(values_only=True):
                val = str(row[0] or "").strip()
                if not val:
                    continue
                if val.startswith("#"):
                    continue
                val_up = val.upper()
                if first_row:
                    first_row = False
                    if val_up in ("FIELD", "FIELD_NAME", "FIELDS", "SAP_FIELD",
                                  "FIELDNAME", "SAP FIELD", "FIELD NAME"):
                        continue
                fields.append(val_up)
            wb.close()

        else:  # CSV
            with open(str(path), encoding="utf-8-sig") as f:
                reader    = csv_mod.reader(f)
                first_row = True
                for row in reader:
                    if not row:
                        continue
                    val = row[0].strip()
                    if not val:
                        continue
                    # Skip comment lines
                    if val.startswith("#"):
                        continue
                    val_up = val.upper()
                    # Skip header row
                    if first_row:
                        first_row = False
                        if val_up in ("FIELD", "FIELD_NAME", "FIELDS", "SAP_FIELD",
                                      "FIELDNAME", "SAP FIELD", "FIELD NAME"):
                            continue
                    fields.append(val_up)

    except Exception as e:
        print(f"Template parse error ({path.name}): {e}")
    return fields


def _read_mapping_file(path: Path) -> dict:
    """
    Read a custom mapping CSV: SOURCE_FIELD,TARGET_FIELD
    Returns {SOURCE: TARGET} dict (uppercase keys).
    Lines starting with # are comments and are ignored.
    """
    mapping = {}
    if not path.exists():
        return mapping
    try:
        with open(str(path), encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            first  = True
            for row in reader:
                if not row or not row[0].strip():
                    continue
                if row[0].strip().startswith("#"):
                    continue
                src = row[0].strip().upper()
                tgt = row[1].strip().upper() if len(row) > 1 else ""
                # Skip header row
                if first:
                    first = False
                    if src in ("SOURCE_FIELD", "SOURCE", "FROM", "SRC"):
                        continue
                if src and tgt:
                    mapping[src] = tgt
    except Exception as e:
        print(f"Mapping file read error: {e}")
    return mapping


def _list_templates() -> list:
    cfg    = load_config()
    active = cfg.get("active_template", "")
    result = []
    for p in sorted(TEMPLATES_DIR.iterdir()):
        if p.suffix.lower() in TEMPLATE_EXT and p.is_file():
            fields = _read_template_fields(p)
            result.append({
                "filename":    p.name,
                "field_count": len(fields),
                "fields":      fields,
                "is_active":   p.name == active,
                "modified":    datetime.fromtimestamp(
                    p.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return result


# ── File discovery & pairing ────────────────────────────────────────────────────

def get_available_files():
    src_dir, tgt_dir = get_dirs()
    src = sorted([f for f in src_dir.iterdir() if f.suffix.lower() in SUPPORTED_EXT],
                 key=lambda f: f.name.upper())
    tgt = sorted([f for f in tgt_dir.iterdir() if f.suffix.lower() in SUPPORTED_EXT],
                 key=lambda f: f.name.upper())
    return src, tgt


def discover_pairs():
    SOURCE_DIR, TARGET_DIR = get_dirs()
    cfg          = load_config()
    manual_pairs = cfg.get("manual_pairs", [])

    src_files = {f.name: f for f in SOURCE_DIR.iterdir()
                 if f.suffix.lower() in SUPPORTED_EXT}
    tgt_files = {f.name: f for f in TARGET_DIR.iterdir()
                 if f.suffix.lower() in SUPPORTED_EXT}

    pairs    = []
    used_src = set()
    used_tgt = set()

    # Manual pairs first
    for mp in manual_pairs:
        src_name = mp.get("source_file", "")
        tgt_name = mp.get("target_file", "")
        name     = mp.get("name", "").upper().strip() or Path(src_name).stem.upper()
        sp       = str(src_files[src_name]) if src_name in src_files else None
        tp       = str(tgt_files[tgt_name]) if tgt_name in tgt_files else None
        has_pair = bool(sp and tp)
        mtime    = max(Path(sp).stat().st_mtime, Path(tp).stat().st_mtime) if has_pair else None
        pairs.append({
            "name": name, "source_path": sp, "target_path": tp,
            "has_pair": has_pair, "mtime": mtime,
            "source_file": Path(sp).name if sp else src_name,
            "target_file": Path(tp).name if tp else tgt_name,
            "match_type": "manual",
            "missing": [] if has_pair else
                       (["source"] if not sp else []) + (["target"] if not tp else []),
        })
        if sp: used_src.add(src_name)
        if tp: used_tgt.add(tgt_name)

    # Auto-pair by exact filename stem
    src_by_stem = {Path(f).stem.upper(): (f, fp)
                   for f, fp in src_files.items() if f not in used_src}
    tgt_by_stem = {Path(f).stem.upper(): (f, fp)
                   for f, fp in tgt_files.items() if f not in used_tgt}

    for stem in sorted(set(src_by_stem) & set(tgt_by_stem)):
        sf, sp = src_by_stem[stem]
        tf, tp = tgt_by_stem[stem]
        mtime  = max(sp.stat().st_mtime, tp.stat().st_mtime)
        pairs.append({
            "name": stem, "source_path": str(sp), "target_path": str(tp),
            "has_pair": True, "mtime": mtime,
            "source_file": sf, "target_file": tf,
            "match_type": "auto", "missing": [],
        })
        used_src.add(sf)
        used_tgt.add(tf)

    # Unmatched
    for f, fp in src_files.items():
        if f not in used_src:
            pairs.append({"name": Path(f).stem.upper(), "source_path": str(fp),
                          "target_path": None, "has_pair": False, "mtime": None,
                          "source_file": f, "target_file": None,
                          "match_type": "unmatched", "missing": ["target"]})
    for f, fp in tgt_files.items():
        if f not in used_tgt:
            pairs.append({"name": Path(f).stem.upper(), "source_path": None,
                          "target_path": str(fp), "has_pair": False, "mtime": None,
                          "source_file": None, "target_file": f,
                          "match_type": "unmatched", "missing": ["source"]})
    return pairs


# ── Business status ─────────────────────────────────────────────────────────────

def calculate_business_status(result, pass_threshold):
    ss        = result.summary_stats
    pass_rate = float(ss.get("pass_rate_pct", 0))
    only_src  = int(result.records_only_in_source or 0)
    only_tgt  = int(result.records_only_in_target or 0)

    if pass_rate < pass_threshold:
        return {"status": "FAIL", "field_status": "FAIL", "record_status": "CHECKED",
                "message": (f"Field validation failed. Pass rate is {pass_rate:.2f}% "
                            f"which is below threshold {pass_threshold:.2f}%.")}

    if only_src > 0 or only_tgt > 0:
        if only_src > 0 and only_tgt > 0:
            msg = (f"{only_src:,} records only in source and "
                   f"{only_tgt:,} records only in target.")
        elif only_tgt > 0:
            msg = f"Target has {only_tgt:,} extra records not in source."
        else:
            msg = f"Source has {only_src:,} records not found in target."
        return {"status": "WARNING", "field_status": "PASS", "record_status": "WARNING",
                "message": (f"Field validation passed ({pass_rate:.2f}% >= "
                            f"{pass_threshold:.2f}%), but {msg}")}

    return {"status": "PASS", "field_status": "PASS", "record_status": "PASS",
            "message": (f"Validation passed. Field pass rate {pass_rate:.2f}% "
                        f"and records fully reconciled.")}


# ── Core validation runner ──────────────────────────────────────────────────────

def run_validation(name, source_path, target_path):
    cfg            = load_config()
    pass_threshold = float(cfg.get("pass_threshold", 100.0))
    custom         = _get_custom_labels()

    # ── Determine field filter: template > manual > all ───────────────────────
    selected_fields    = cfg.get("selected_fields", [])
    active_template    = cfg.get("active_template", "")
    template_name_used = ""

    if active_template:
        tmpl_path = TEMPLATES_DIR / active_template
        if tmpl_path.exists():
            tmpl_fields = _read_template_fields(tmpl_path)
            if tmpl_fields:
                selected_fields    = tmpl_fields
                template_name_used = active_template
                log_event(f"{name}: template '{active_template}' ({len(tmpl_fields)} fields)", "info")
        else:
            log_event(f"{name}: template '{active_template}' not found", "warn")

    # ── Manual join keys: from config (user set via UI) ───────────────────────
    manual_join_keys = cfg.get("manual_join_keys", {}).get(name.upper(), [])

    src_mb = Path(source_path).stat().st_size / (1024 * 1024)
    tgt_mb = Path(target_path).stat().st_size / (1024 * 1024)
    if src_mb > 50 or tgt_mb > 50:
        log_event(f"{name}: large files ({src_mb:.1f}MB / {tgt_mb:.1f}MB)", "warn")

    # ── Read headers ──────────────────────────────────────────────────────────
    try:
        src_cols, tgt_cols = _read_file_headers(source_path, target_path)
    except Exception as e:
        log_event(f"{name}: could not read headers — {e}", "warn")
        src_cols, tgt_cols = [], []

    # ── Build field mapping ───────────────────────────────────────────────────
    # Target columns are authoritative; filter by template/selection if active
    tgt_no_jk = tgt_cols  # key cols stripped later by validator
    if selected_fields:
        sel_upper    = set(s.upper() for s in selected_fields)
        tgt_filtered = [c for c in tgt_no_jk if c in sel_upper] or tgt_no_jk
    else:
        tgt_filtered = tgt_no_jk

    src_no_jk = src_cols

    custom_mapping = _read_mapping_file(MAPPING_FILE)
    if custom_mapping:
        src_set   = set(src_no_jk)
        tgt_set   = set(tgt_filtered)
        field_map = {s: t for s, t in custom_mapping.items()
                     if s in src_set and t in tgt_set}
        if field_map:
            log_event(f"{name}: custom mapping — {len(field_map)} pairs", "info")
            from core.field_mapper import MappingResult, MappedField
            mapping_result = MappingResult(
                mapped_fields=field_map,
                mapped_details=[MappedField(source_field=s, target_field=t,
                    method="custom", confidence=1.0,
                    source_label=get_label(s, custom), target_label=get_label(t, custom))
                    for s, t in field_map.items()],
                unmapped_source=[c for c in src_no_jk if c not in field_map],
                unmapped_target=[c for c in tgt_filtered if c not in field_map.values()],
                suggested_mappings=[], object_type=name,
                total_source_fields=len(src_no_jk), total_target_fields=len(tgt_filtered),
            )
        else:
            custom_mapping = {}

    if not custom_mapping:
        mapping_result = build_field_mapping(
            source_cols=src_no_jk, target_cols=tgt_filtered,
            object_type=name, selected_fields=None, custom_labels=custom,
        )
        field_map = mapping_result.mapped_fields

    exact_count = sum(1 for d in mapping_result.mapped_details if d.method == "exact")
    alias_count = sum(1 for d in mapping_result.mapped_details if "alias" in d.method)
    fuzzy_count = sum(1 for d in mapping_result.mapped_details if d.method == "fuzzy")
    log_event(
        f"{name}: {len(field_map)} fields mapped "
        f"({exact_count} exact, {alias_count} alias, {fuzzy_count} fuzzy)"
        + (f" via template '{template_name_used}'" if template_name_used else ""),
        "info",
    )

    # ── Run validator (composite key, column-only loading) ────────────────────
    validator = MaterialValidator(
        field_map=field_map,
        pass_threshold=pass_threshold,
        manual_join_keys=manual_join_keys if manual_join_keys else None,
        custom_labels=custom if custom else None,
    )

    result          = validator.validate(source_path, target_path, object_name=name)
    ss              = result.summary_stats
    business_status = calculate_business_status(result, pass_threshold)

    # Log join key info
    jk_str = " + ".join(result.join_keys) if result.join_keys else "none"
    log_event(
        f"{name}: join keys = [{jk_str}] "
        f"method={result.key_detection_method} "
        f"confidence={result.key_confidence} "
        f"dup_src={result.duplicate_src} dup_tgt={result.duplicate_tgt}",
        "info" if result.key_confidence in ("high","medium") else "warn",
    )

    # ── Build field rows ───────────────────────────────────────────────────────
    field_rows = []
    for fr in result.field_results:
        detail = next(
            (d for d in mapping_result.mapped_details if d.source_field == fr.field_source),
            None,
        )
        disp = get_display(fr.field_source, fr.field_target, custom)
        field_rows.append({
            "field":              fr.field_source,
            "field_label":        disp["source_label"],
            "field_target":       fr.field_target,
            "field_target_label": disp["target_label"],
            "display_name":       disp["display_name"],
            "display_mapping":    disp["display_mapping"],
            "is_cross_mapped":    disp["is_cross_mapped"],
            "is_key_field":       fr.is_key_field,          # join key field flag
            "mapping_method":     detail.method if detail else ("key" if fr.is_key_field else "exact"),
            "mapping_confidence": detail.confidence if detail else 1.0,
            "type":               "numeric" if fr.is_numeric else "string",
            "tolerance":          fr.tolerance_used,
            "total":              fr.total_records,
            "matched":            fr.matched,
            "mismatched":         fr.mismatched,
            "miss_source":        fr.missing_in_source,
            "miss_target":        fr.missing_in_target,
            "match_pct":          fr.match_pct,
            "pass_threshold":     fr.pass_threshold,
            "status":             fr.status,
            "mismatches":         fr.mismatch_details,
            "matches":            fr.matched_details,
            "mismatch_count":     len(fr.mismatch_details),
        })

    # ── Mapping info for dashboard ─────────────────────────────────────────────
    mapping_info = None
    if result.mapping:
        m = result.mapping
        mapping_info = {
            # Composite key fields
            "join_keys":          result.join_keys,
            "join_key_labels":    {k: get_label(k, custom) for k in result.join_keys},
            "join_key":           m.join_key,           # backwards compat string
            "join_key_label":     m.join_key_label,
            "key_detection_method": result.key_detection_method,
            "key_confidence":     result.key_confidence,
            "duplicate_src":      result.duplicate_src,
            "duplicate_tgt":      result.duplicate_tgt,
            "duplicate_key_samples": result.duplicate_key_samples,
            # Field info
            "matched_fields":     m.matched_fields,
            "matched_labels":     {f: get_label(f, custom) for f in m.matched_fields},
            "source_only_fields": mapping_result.unmapped_source,
            "source_only_labels": {f: get_label(f, custom) for f in mapping_result.unmapped_source},
            "target_only_fields": mapping_result.unmapped_target,
            "target_only_labels": {f: get_label(f, custom) for f in mapping_result.unmapped_target},
            "numeric_fields":     m.numeric_fields,
            "tolerance_map":      m.tolerance_map,
            "selected_fields":    selected_fields,
            "pass_threshold":     pass_threshold,
        }

    # ── Available fields for Settings ─────────────────────────────────────────
    sel_set = set(selected_fields) if selected_fields else set()
    available_fields = []
    for col in src_cols:
        tgt_col = field_map.get(col)
        available_fields.append({
            "field": col, "label": get_label(col, custom),
            "in_source": True, "in_target": tgt_col is not None,
            "target_col": tgt_col or "", "target_label": get_label(tgt_col, custom) if tgt_col else "",
            "common": tgt_col is not None, "selected": not sel_set or col in sel_set,
        })
    for col in tgt_cols:
        if col not in field_map.values():
            available_fields.append({
                "field": col, "label": get_label(col, custom),
                "in_source": False, "in_target": True, "target_col": col,
                "target_label": get_label(col, custom), "common": False, "selected": False,
            })

    ts             = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{name}_{ts}.xlsx"
    excel_path     = REPORTS_DIR / excel_filename

    result_dict = {
        "name":                   name,
        "sap_object":             get_object_config(name).get("description", name),
        "status":                 business_status["status"],
        "validator_status":       result.overall_status,
        "field_status":           business_status["field_status"],
        "record_status":          business_status["record_status"],
        "business_message":       business_status["message"],
        "source_file":            Path(source_path).name,
        "target_file":            Path(target_path).name,
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "pass_threshold":         pass_threshold,
        "selected_fields":        selected_fields,
        "template_used":          template_name_used,
        # Composite key
        "join_keys":              result.join_keys,
        "key_detection_method":   result.key_detection_method,
        "key_confidence":         result.key_confidence,
        "duplicate_src":          result.duplicate_src,
        "duplicate_tgt":          result.duplicate_tgt,
        "duplicate_key_samples":  result.duplicate_key_samples,
        "manual_join_keys":       manual_join_keys,
        "errors":                 result.errors,
        "mapping":                mapping_info,
        "field_mapping_detail":   mapping_result_to_dict(mapping_result),
        "field_results":          field_rows,
        "available_fields":       available_fields,
        "run_at":                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "excel_file":             excel_filename,
        "corrected_files":        find_corrected_files(Path(target_path).name),
    }
    result_dict["recommendations"] = build_recommendations(result_dict)

    try:
        generate_excel_report(result_dict, str(excel_path))
        cleanup_old_reports()
    except Exception as e:
        result_dict["excel_error"] = str(e)
        log_event(f"Excel failed for {name}: {e}", "error")

    cache_and_trim_row_details(result_dict)
    return result_dict
    cfg            = load_config()
    pass_threshold = float(cfg.get("pass_threshold", 100.0))
    custom         = _get_custom_labels()

    # ── Determine field filter: template > manual > all ───────────────────────
    selected_fields    = cfg.get("selected_fields", [])
    active_template    = cfg.get("active_template", "")
    template_name_used = ""

    if active_template:
        tmpl_path = TEMPLATES_DIR / active_template
        if tmpl_path.exists():
            tmpl_fields = _read_template_fields(tmpl_path)
            if tmpl_fields:
                selected_fields    = tmpl_fields
                template_name_used = active_template
                log_event(
                    f"{name}: using template '{active_template}' "
                    f"({len(tmpl_fields)} fields)",
                    "info",
                )
            else:
                log_event(
                    f"{name}: template '{active_template}' is empty — "
                    f"validating all fields",
                    "warn",
                )
        else:
            log_event(
                f"{name}: template '{active_template}' not found — "
                f"validating all fields",
                "warn",
            )

    obj_cfg  = get_object_config(name)
    join_key = obj_cfg.get("join_key", None)

    src_mb = Path(source_path).stat().st_size / (1024 * 1024)
    tgt_mb = Path(target_path).stat().st_size / (1024 * 1024)
    if src_mb > 50 or tgt_mb > 50:
        log_event(
            f"{name}: large files ({src_mb:.1f} MB / {tgt_mb:.1f} MB) — "
            f"may take a few minutes",
            "warn",
        )

    try:
        src_cols, tgt_cols = _read_file_headers(source_path, target_path)
    except Exception as e:
        log_event(f"{name}: could not read headers — {e}", "warn")
        src_cols, tgt_cols = [], []

    jk_upper  = join_key.upper() if join_key else ""
    src_no_jk = [c for c in src_cols if c != jk_upper]
    tgt_no_jk = [c for c in tgt_cols if c != jk_upper]

    # ── Build field mapping: source → target ──────────────────────────────────
    # Forward approach: for each source column find its target equivalent.
    # The target file drives WHAT gets validated — we use all target columns
    # as the reference set. selected_fields/template filters which target
    # columns we care about.

    if selected_fields:
        sel_upper = set(s.upper() for s in selected_fields)
        # Filter to target columns the user selected
        tgt_filtered = [c for c in tgt_no_jk if c in sel_upper]
        if not tgt_filtered:
            tgt_filtered = tgt_no_jk  # fallback: all target columns
        log_event(
            f"{name}: filtering to {len(tgt_filtered)} of "
            f"{len(tgt_no_jk)} target columns",
            "info",
        )
    else:
        tgt_filtered = tgt_no_jk

    # ── Build field mapping: custom file > alias > fuzzy ─────────────────────
    # If user has uploaded a custom_mapping.csv, use it directly.
    # Otherwise use the alias+fuzzy engine.
    custom_mapping = _read_mapping_file(MAPPING_FILE)
    if custom_mapping:
        # Filter to pairs where both columns exist
        src_set_check = set(src_no_jk)
        tgt_set_check = set(tgt_filtered)
        filtered_map  = {}
        for s, t in custom_mapping.items():
            if s in src_set_check and t in tgt_set_check:
                filtered_map[s] = t
        if filtered_map:
            log_event(
                f"{name}: using custom mapping file — "
                f"{len(filtered_map)} of {len(custom_mapping)} pairs matched",
                "info",
            )
            # Build a minimal MappingResult from the custom map
            from core.field_mapper import MappingResult, MappedField
            details = [
                MappedField(
                    source_field=s, target_field=t,
                    method="custom", confidence=1.0,
                    source_label=get_label(s, custom),
                    target_label=get_label(t, custom),
                )
                for s, t in filtered_map.items()
            ]
            mapping_result = MappingResult(
                mapped_fields=filtered_map,
                mapped_details=details,
                unmapped_source=[c for c in src_no_jk if c not in filtered_map],
                unmapped_target=[c for c in tgt_filtered if c not in filtered_map.values()],
                suggested_mappings=[],
                object_type=name,
                total_source_fields=len(src_no_jk),
                total_target_fields=len(tgt_filtered),
            )
        else:
            log_event(
                f"{name}: custom mapping file has no matching pairs for this file — "
                f"falling back to auto mapping",
                "warn",
            )
            custom_mapping = {}

    if not custom_mapping:
        mapping_result = build_field_mapping(
            source_cols=src_no_jk,
            target_cols=tgt_filtered,
            object_type=name,
            selected_fields=None,
            custom_labels=custom,
        )

    field_map    = mapping_result.mapped_fields
    exact_count  = sum(1 for d in mapping_result.mapped_details if d.method == "exact")
    alias_count  = sum(1 for d in mapping_result.mapped_details if "alias" in d.method)
    fuzzy_count  = sum(1 for d in mapping_result.mapped_details if d.method == "fuzzy")

    log_event(
        f"{name}: mapped {len(field_map)} fields "
        f"({exact_count} exact, {alias_count} alias, {fuzzy_count} fuzzy)"
        + (f" via template '{template_name_used}'" if template_name_used else ""),
        "info",
    )

    # Warn if template produced zero mapped fields
    if selected_fields and template_name_used and not field_map:
        log_event(
            f"{name}: WARNING — template '{template_name_used}' has "
            f"{len(selected_fields)} fields but NONE matched any source column. "
            f"Source columns: {src_no_jk[:8]}... "
            f"Template fields: {selected_fields[:8]}...",
            "error",
        )
    elif selected_fields and template_name_used:
        # Log which template fields weren't found
        mapped_set = set(field_map.keys())
        src_set    = set(src_no_jk)
        missing    = [f for f in selected_fields
                      if f not in src_set and f not in mapped_set]
        if missing:
            log_event(
                f"{name}: {len(missing)} template field(s) not matched: "
                f"{', '.join(missing[:8])}"
                + (" …" if len(missing) > 8 else ""),
                "warn",
            )

    validator = MaterialValidator(
        field_map=field_map,
        pass_threshold=pass_threshold,
        join_key=join_key,
        custom_labels=custom if custom else None,
    )

    result          = validator.validate(source_path, target_path)
    ss              = result.summary_stats
    business_status = calculate_business_status(result, pass_threshold)

    # ── Build field rows ───────────────────────────────────────────────────────
    field_rows = []
    for fr in result.field_results:
        detail = next(
            (d for d in mapping_result.mapped_details
             if d.source_field == fr.field_source), None
        )
        disp = get_display(fr.field_source, fr.field_target, custom)
        field_rows.append({
            "field":              fr.field_source,
            "field_label":        disp["source_label"],
            "field_target":       fr.field_target,
            "field_target_label": disp["target_label"],
            "display_name":       disp["display_name"],
            "display_mapping":    disp["display_mapping"],
            "is_cross_mapped":    disp["is_cross_mapped"],
            "mapping_method":     detail.method if detail else "exact",
            "mapping_confidence": detail.confidence if detail else 1.0,
            "type":               "numeric" if fr.is_numeric else "string",
            "tolerance":          fr.tolerance_used,
            "total":              fr.total_records,
            "matched":            fr.matched,
            "mismatched":         fr.mismatched,
            "miss_source":        fr.missing_in_source,
            "miss_target":        fr.missing_in_target,
            "match_pct":          fr.match_pct,
            "pass_threshold":     fr.pass_threshold,
            "status":             fr.status,
            "mismatches":         fr.mismatch_details,
            "matches":            fr.matched_details,
            "mismatch_count":     len(fr.mismatch_details),
        })

    # ── Mapping info for dashboard ─────────────────────────────────────────────
    mapping_info = None
    if result.mapping:
        mapping_info = {
            "join_key":           result.mapping.join_key,
            "join_key_label":     get_label(result.mapping.join_key, custom),
            "matched_fields":     result.mapping.matched_fields,
            "matched_labels":     {f: get_label(f, custom)
                                   for f in result.mapping.matched_fields},
            "source_only_fields": mapping_result.unmapped_source,
            "source_only_labels": {f: get_label(f, custom)
                                   for f in mapping_result.unmapped_source},
            "target_only_fields": mapping_result.unmapped_target,
            "target_only_labels": {f: get_label(f, custom)
                                   for f in mapping_result.unmapped_target},
            "numeric_fields":     result.mapping.numeric_fields,
            "tolerance_map":      result.mapping.tolerance_map,
            "selected_fields":    selected_fields,
            "pass_threshold":     pass_threshold,
        }

    # ── available_fields for Settings field selector ───────────────────────────
    sel_set = set(selected_fields) if selected_fields else set()
    available_fields = []
    for col in src_cols:
        tgt_col = field_map.get(col)
        available_fields.append({
            "field":        col,
            "label":        get_label(col, custom),
            "in_source":    True,
            "in_target":    tgt_col is not None,
            "target_col":   tgt_col or "",
            "target_label": get_label(tgt_col, custom) if tgt_col else "",
            "common":       tgt_col is not None,
            "selected":     not sel_set or col in sel_set,
        })
    for col in tgt_cols:
        if col not in field_map.values() and col != jk_upper:
            available_fields.append({
                "field":        col,
                "label":        get_label(col, custom),
                "in_source":    False,
                "in_target":    True,
                "target_col":   col,
                "target_label": get_label(col, custom),
                "common":       False,
                "selected":     False,
            })

    ts             = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{name}_{ts}.xlsx"
    excel_path     = REPORTS_DIR / excel_filename

    result_dict = {
        "name":                   name,
        "sap_object":             obj_cfg.get("description", name),
        "status":                 business_status["status"],
        "validator_status":       result.overall_status,
        "field_status":           business_status["field_status"],
        "record_status":          business_status["record_status"],
        "business_message":       business_status["message"],
        "source_file":            Path(source_path).name,
        "target_file":            Path(target_path).name,
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "pass_threshold":         pass_threshold,
        "selected_fields":        selected_fields,
        "template_used":          template_name_used,
        "errors":                 result.errors,
        "mapping":                mapping_info,
        "field_mapping_detail":   mapping_result_to_dict(mapping_result),
        "field_results":          field_rows,
        "available_fields":       available_fields,
        "run_at":                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "excel_file":             excel_filename,
        "corrected_files":        find_corrected_files(Path(target_path).name),
    }

    result_dict["recommendations"] = build_recommendations(result_dict)

    try:
        generate_excel_report(result_dict, str(excel_path))
        cleanup_old_reports()
    except Exception as e:
        result_dict["excel_error"] = str(e)
        log_event(f"Excel failed for {name}: {e}", "error")

    cache_and_trim_row_details(result_dict)
    return result_dict


# ── Scan orchestrator ───────────────────────────────────────────────────────────

def scan_and_validate_all():
    if not scan_lock.acquire(blocking=False):
        log_event("Scan already running — skipping", "warn")
        return

    scan_status.update({
        "scanning": True, "error": None,
        "current_file": None, "total_files": 0, "completed_files": 0,
    })
    try:
        pairs       = discover_pairs()
        valid_pairs = [p for p in pairs if p["has_pair"]]
        scan_status["total_files"] = len(valid_pairs)

        for pair in pairs:
            name = pair["name"]

            if not pair["has_pair"]:
                if file_states.get(name, {}).get("state") != "unmatched":
                    side  = "source" if pair["source_path"] else "target"
                    other = "target" if side == "source" else "source"
                    log_event(
                        f"{name}: found in {side} only — "
                        f"waiting for {other} to pair",
                        "warn",
                    )
                    file_states[name] = {
                        "state": "unmatched",
                        "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "source_file": pair["source_file"],
                        "target_file": pair["target_file"],
                    }
                continue

            last_mtime = pair["mtime"]
            existing   = results_store.get(name)
            prev_state = file_states.get(name, {})

            if not existing:
                log_event(
                    f"{name}: new pair [{pair.get('match_type','auto')}] — "
                    f"{pair['source_file']} ↔ {pair['target_file']}",
                    "info",
                )
            elif prev_state.get("_mtime") != last_mtime:
                log_event(f"{name}: file changed — re-validating", "info")
            else:
                scan_status["completed_files"] += 1
                continue

            scan_status["current_file"] = name
            file_states[name] = {
                "state": "validating",
                "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_file": pair["source_file"],
                "target_file": pair["target_file"],
                "_mtime": last_mtime,
            }

            try:
                result = run_validation(
                    name, pair["source_path"], pair["target_path"]
                )
                result["_mtime"] = last_mtime
                results_store[name] = result

                file_states[name] = {
                    "state":         "done",
                    "detected_at":   file_states[name]["detected_at"],
                    "validated_at":  result["run_at"],
                    "source_file":   pair["source_file"],
                    "target_file":   pair["target_file"],
                    "_mtime":        last_mtime,
                    "status":        result["status"],
                    "field_status":  result["field_status"],
                    "record_status": result["record_status"],
                    "message":       result["business_message"],
                }

                level = ("success" if result["status"] == "PASS"
                         else "warn" if result["status"] == "WARNING"
                         else "error")
                log_event(
                    f"{name}: {result['status']} — {result['business_message']} | "
                    f"Matched: {result['records_matched']:,} | "
                    f"Src only: {result['records_only_in_source']:,} | "
                    f"Tgt only: {result['records_only_in_target']:,}",
                    level,
                )
            except Exception as e:
                file_states[name]["state"] = "error"
                file_states[name]["error"] = str(e)
                scan_status["error"]       = str(e)
                log_event(f"{name}: ERROR — {e}", "error")
            finally:
                scan_status["completed_files"] += 1

        scan_status["last_scan"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    except Exception as e:
        scan_status["error"] = str(e)
        log_event(f"Scan error: {e}", "error")
    finally:
        scan_status["scanning"]     = False
        scan_status["current_file"] = None
        scan_lock.release()


def background_watcher(interval=60):
    while True:
        scan_and_validate_all()
        time.sleep(interval)


# ── Routes ──────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("dashboard.html")


@app.route("/api/scan", methods=["POST"])
def api_scan():
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/status")
def api_status():
    pairs = discover_pairs()
    cfg   = load_config()
    s, t  = get_dirs()
    sel   = cfg.get("selected_fields", [])
    tmpl  = cfg.get("active_template", "")
    return jsonify({
        "last_scan":       scan_status["last_scan"],
        "scanning":        scan_status["scanning"],
        "error":           scan_status["error"],
        "current_file":    scan_status["current_file"],
        "total_files":     scan_status["total_files"],
        "completed_files": scan_status["completed_files"],
        "source_dir":      str(s),
        "target_dir":      str(t),
        "pairs":           pairs,
        "file_states":     file_states,
        "total_tables":    len([p for p in pairs if p["has_pair"]]),
        "unmatched":       len([p for p in pairs if not p["has_pair"]]),
        "pass_threshold":  cfg.get("pass_threshold", 100.0),
        "selected_fields": sel,
        "active_template": tmpl,
        "validation_mode": (
            f"template:{tmpl}" if tmpl else
            "selected_fields"  if sel  else
            "all_fields"
        ),
    })


@app.route("/api/results")
def api_results():
    return jsonify(list(results_store.values()))


@app.route("/api/results/<name>")
def api_result_detail(name):
    r = results_store.get(name.upper())
    return jsonify(r) if r else (jsonify({"error": "Not found"}), 404)


@app.route("/field-records/<name>/<path:field>")
def field_records_page(name, field):
    """Paginated matched-record drill-down opened by double-clicking a PASS field."""
    details = row_details_store.get(name.upper(), {}).get(field.upper())
    if not details:
        return Response("Matched-record details are not available. Run validation again.", status=404)
    rows = details.get("matches", [])
    page_size = 500
    try: page = max(1, int(request.args.get("page", 1)))
    except ValueError: page = 1
    pages = max(1, (len(rows) + page_size - 1) // page_size)
    page = min(page, pages)
    start = (page - 1) * page_size
    visible = rows[start:start + page_size]
    esc_html = lambda value: html_lib.escape(str(value if value is not None else ""))
    body_rows = "".join(
        "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(
            esc_html(row.get("material", "")), esc_html(row.get("source_value", "")),
            esc_html(row.get("target_value", "")), esc_html(row.get("result", "")))
        for row in visible
    )
    base = f"/field-records/{quote(name, safe='')}/{quote(field, safe='')}"
    nav = ""
    if page > 1: nav += f'<a href="{base}?page={page-1}">Previous</a>'
    nav += f"<span>Page {page} of {pages} · {len(rows):,} matched records</span>"
    if page < pages: nav += f'<a href="{base}?page={page+1}">Next</a>'
    document = f"""<!doctype html><html><head><meta charset="utf-8">
    <title>Matched Records - {esc_html(details.get('label', field))}</title>
    <style>body{{font:13px system-ui;margin:24px;background:#f4f6fa;color:#1a1f36}}
    h1{{font-size:20px}}p{{color:#6b728e}}table{{width:100%;border-collapse:collapse;background:white}}
    th{{background:#134e4a;color:white;text-align:left;padding:9px}}td{{padding:8px;border-bottom:1px solid #dde1ec}}
    tr:nth-child(even){{background:#f0fdf4}}.nav{{display:flex;gap:16px;align-items:center;margin:16px 0}}
    a{{color:#0f766e;font-weight:700}}</style></head><body>
    <h1>{esc_html(details.get('label', field))} — Matched Records</h1>
    <p>{esc_html(name)} · Source field {esc_html(field)} · Target field {esc_html(details.get('target_field', field))}</p>
    <div class="nav">{nav}</div><table><thead><tr><th>Key</th><th>Source Value</th><th>Target Value</th><th>Result</th></tr></thead>
    <tbody>{body_rows}</tbody></table><div class="nav">{nav}</div></body></html>"""
    return Response(document, mimetype="text/html")


@app.route("/field-errors/<name>/<path:field>")
def field_errors_page(name, field):
    """Paginated error drill-down opened by double-clicking a FAIL field."""
    details = row_details_store.get(name.upper(), {}).get(field.upper())
    if not details:
        return Response("Error details are not available. Run validation again.", status=404)
    rows = details.get("mismatches", [])
    page_size = 500
    try: page = max(1, int(request.args.get("page", 1)))
    except ValueError: page = 1
    pages = max(1, (len(rows) + page_size - 1) // page_size)
    page = min(page, pages)
    start = (page - 1) * page_size
    visible = rows[start:start + page_size]
    esc_html = lambda value: html_lib.escape(str(value if value is not None else ""))
    body_rows = "".join(
        "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(
            esc_html(row.get("material", "")), esc_html(row.get("source_value", "")),
            esc_html(row.get("target_value", "")), esc_html(row.get("issue", "")))
        for row in visible
    )
    base = f"/field-errors/{quote(name, safe='')}/{quote(field, safe='')}"
    nav = ""
    if page > 1: nav += f'<a href="{base}?page={page-1}">Previous</a>'
    nav += f"<span>Page {page} of {pages} · {len(rows):,} error records</span>"
    if page < pages: nav += f'<a href="{base}?page={page+1}">Next</a>'
    document = f"""<!doctype html><html><head><meta charset="utf-8">
    <title>Error Records - {esc_html(details.get('label', field))}</title>
    <style>body{{font:13px system-ui;margin:24px;background:#fff7f6;color:#1a1f36}}
    h1{{font-size:20px}}p{{color:#6b728e}}table{{width:100%;border-collapse:collapse;background:white}}
    th{{background:#b42318;color:white;text-align:left;padding:9px}}td{{padding:8px;border-bottom:1px solid #f0d5d2}}
    tr:nth-child(even){{background:#fff1f0}}.nav{{display:flex;gap:16px;align-items:center;margin:16px 0}}
    a{{color:#b42318;font-weight:700}}</style></head><body>
    <h1>{esc_html(details.get('label', field))} — Error Records</h1>
    <p>{esc_html(name)} · Source field {esc_html(field)} · Target field {esc_html(details.get('target_field', field))}</p>
    <div class="nav">{nav}</div><table><thead><tr><th>Key</th><th>Source Value</th><th>Target Value</th><th>Issue</th></tr></thead>
    <tbody>{body_rows}</tbody></table><div class="nav">{nav}</div></body></html>"""
    return Response(document, mimetype="text/html")


@app.route("/api/activity")
def api_activity():
    return jsonify(list(reversed(activity_log)))


# ── Upload ──────────────────────────────────────────────────────────────────────

@app.route("/api/upload/source", methods=["POST"])
def upload_source():
    return _handle_upload(request, get_dirs()[0], "source")


@app.route("/api/upload/target", methods=["POST"])
def upload_target():
    return _handle_upload(request, get_dirs()[1], "target")


def _handle_upload(req, dest_dir, side):
    if "file" not in req.files:
        return jsonify({"error": "No file part"}), 400
    saved, errors = [], []
    for f in req.files.getlist("file"):
        if not f.filename:
            continue
        save_name = secure_filename(f.filename)
        if Path(save_name).suffix.lower() not in SUPPORTED_EXT:
            errors.append(f"Unsupported type: {save_name}")
            continue
        f.save(str(dest_dir / save_name))
        log_event(f"Uploaded to {side}: {save_name}", "info")
        saved.append(save_name)
    if saved:
        threading.Thread(target=scan_and_validate_all, daemon=True).start()
    if errors and not saved:
        return jsonify({"error": "; ".join(errors)}), 400

    # ── Read headers immediately from uploaded file so dashboard can
    # populate the field selector without waiting for a scan ──────────────────
    custom   = _get_custom_labels()
    headers  = {}
    for fname in saved:
        fpath = dest_dir / fname
        try:
            if side == "source":
                cols, _ = _read_file_headers(str(fpath))
            else:
                _, cols = _read_file_headers(None, str(fpath))
            cols = cols or []
            headers[fname] = {
                "columns": cols,
                "labels":  {c: get_label(c, custom) for c in cols},
                "count":   len(cols),
            }
            log_event(
                f"Headers read from {side}/{fname}: "
                f"{len(cols)} columns — {', '.join(cols[:6])}"
                + (" …" if len(cols) > 6 else ""),
                "info",
            )
        except Exception as e:
            headers[fname] = {"error": str(e), "columns": [], "count": 0}
            log_event(f"Could not read headers from {fname}: {e}", "warn")

    return jsonify({"ok": True, "saved": saved, "warnings": errors,
                    "side": side, "headers": headers})


@app.route("/api/upload/labels", methods=["POST"])
def upload_labels():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400
    f.save(str(LABELS_FILE))
    log_event(f"Custom labels uploaded: {secure_filename(f.filename)}", "info")
    results_store.clear()
    for n in file_states:
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/upload/mapping", methods=["POST"])
def upload_mapping():
    """
    Upload a custom field mapping CSV.
    Format: SOURCE_FIELD,TARGET_FIELD  (one pair per row, header optional)
    Example:
        KUNNR,KUNNR
        NAME1,NAMORG1
        LAND1,COUNTRY
        ORT01,CITY1
    This overrides the alias/fuzzy auto-mapping for all validations.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400
    if Path(f.filename).suffix.lower() not in {".csv", ".txt"}:
        return jsonify({"error": "Use a CSV or TXT file"}), 400
    f.save(str(MAPPING_FILE))
    pairs = _read_mapping_file(MAPPING_FILE)
    log_event(
        f"Custom mapping uploaded: {len(pairs)} pairs — "
        f"{', '.join(f'{s}->{t}' for s,t in list(pairs.items())[:5])}"
        + (" …" if len(pairs) > 5 else ""),
        "info",
    )
    results_store.clear()
    for n in file_states:
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True, "pairs": len(pairs), "mapping": pairs})


@app.route("/api/upload/mapping/clear", methods=["POST"])
def clear_mapping():
    """Remove the custom mapping file — revert to auto alias/fuzzy mapping."""
    if MAPPING_FILE.exists():
        MAPPING_FILE.unlink()
        log_event("Custom mapping cleared — using auto alias/fuzzy mapping", "info")
        results_store.clear()
        for n in file_states:
            if file_states[n].get("state") == "done":
                file_states[n]["state"] = "changed"
        threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/upload/mapping/status", methods=["GET"])
def mapping_status():
    """Return current custom mapping pairs if a mapping file is uploaded."""
    if MAPPING_FILE.exists():
        pairs = _read_mapping_file(MAPPING_FILE)
        return jsonify({"active": True, "pairs": len(pairs), "mapping": pairs})
    return jsonify({"active": False, "pairs": 0, "mapping": {}})


@app.route("/api/upload/mapping/sample", methods=["GET"])
def mapping_sample():
    """Download a sample mapping CSV."""
    lines = [
        "SOURCE_FIELD,TARGET_FIELD",
        "# One pair per row. Lines starting with # are ignored.",
        "# Use this when auto-mapping doesn't work correctly.",
        "KUNNR,KUNNR",
        "NAME1,NAMORG1",
        "KTOKD,KTOKD",
        "LAND1,COUNTRY",
        "ORT01,CITY1",
        "PSTLZ,POST_CODE1",
        "STRAS,STREET",
        "REGIO,REGION",
        "ZTERM,ZTERM",
        "WAERS,WAERS",
    ]
    return Response(
        "\n".join(lines).encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=sample_mapping.csv"},
    )


# ── Template routes ─────────────────────────────────────────────────────────────

@app.route("/api/templates", methods=["GET"])
def api_templates_list():
    return jsonify(_list_templates())


@app.route("/api/templates/upload", methods=["POST"])
def api_template_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400
    save_name = secure_filename(f.filename)
    if Path(save_name).suffix.lower() not in TEMPLATE_EXT:
        return jsonify({"error": "Use CSV, XLSX, or TXT"}), 400
    dest = TEMPLATES_DIR / save_name
    f.save(str(dest))
    fields = _read_template_fields(dest)
    log_event(
        f"Template uploaded: {save_name} ({len(fields)} fields): "
        f"{', '.join(fields[:6])}" + (" …" if len(fields) > 6 else ""),
        "info",
    )
    return jsonify({
        "ok": True, "filename": save_name,
        "field_count": len(fields), "fields": fields,
    })


@app.route("/api/templates/<filename>", methods=["DELETE"])
def api_template_delete(filename):
    safe = secure_filename(filename)
    path = TEMPLATES_DIR / safe
    if not path.exists():
        return jsonify({"error": "Not found"}), 404
    path.unlink()
    cfg = load_config()
    if cfg.get("active_template") == safe:
        cfg["active_template"] = ""
        save_config(cfg)
    log_event(f"Template deleted: {safe}", "info")
    return jsonify({"ok": True})


@app.route("/api/templates/activate", methods=["POST"])
def api_template_activate():
    data     = request.get_json(force=True)
    filename = data.get("filename", "").strip()

    if filename:
        safe = secure_filename(filename)
        path = TEMPLATES_DIR / safe
        if not path.exists():
            return jsonify({"error": f"Template not found: {safe}"}), 404
        fields = _read_template_fields(path)
        cfg    = load_config()
        cfg["active_template"] = safe
        save_config(cfg)
        results_store.clear()
        for n in file_states:
            if file_states[n].get("state") == "done":
                file_states[n]["state"] = "changed"
        threading.Thread(target=scan_and_validate_all, daemon=True).start()
        log_event(
            f"Template activated: {safe} — validating "
            f"{len(fields)} fields: "
            f"{', '.join(fields[:6])}" + (" …" if len(fields) > 6 else ""),
            "info",
        )
        return jsonify({
            "ok": True, "active_template": safe,
            "field_count": len(fields), "fields": fields,
        })

    # Deactivate
    cfg = load_config()
    cfg["active_template"] = ""
    save_config(cfg)
    results_store.clear()
    for n in file_states:
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    log_event("Template deactivated — validating all fields", "info")
    return jsonify({"ok": True, "active_template": ""})


@app.route("/api/templates/sample", methods=["GET"])
def api_template_sample():
    """Download a sample field-selection template CSV."""
    lines = [
        "FIELD_NAME",
        "# One SAP field name per row. Lines starting with # are ignored.",
        "# You can use SAP 4.7 names OR S/4HANA names — both work.",
        "# Example for Customer:",
        "KUNNR",
        "NAME1",
        "KTOKD",
        "LAND1",
        "STRAS",
        "ORT01",
        "PSTLZ",
        "REGIO",
        "ZTERM",
        "WAERS",
        "TELF1",
        "SPRAS",
        "ERDAT",
    ]
    return Response(
        "\n".join(lines).encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition":
                 "attachment; filename=sample_field_template.csv"},
    )


# ── Config ──────────────────────────────────────────────────────────────────────

@app.route("/api/config", methods=["GET"])
def api_get_config():
    cfg    = load_config()
    custom = _get_custom_labels()
    sel    = cfg.get("selected_fields", [])
    sel_set = set(sel)

    # ── Build available_fields from TARGET file columns (authoritative) ──────
    # Target file has ALL expected S/4HANA fields — use those as the field list.
    # Source columns are shown as secondary (S badge) so user can see coverage.
    src_dir, tgt_dir = get_dirs()
    available = []

    try:
        src_files = sorted([f for f in src_dir.iterdir()
                            if f.suffix.lower() in SUPPORTED_EXT],
                           key=lambda f: f.stat().st_mtime, reverse=True)
        tgt_files = sorted([f for f in tgt_dir.iterdir()
                            if f.suffix.lower() in SUPPORTED_EXT],
                           key=lambda f: f.stat().st_mtime, reverse=True)

        src_path = str(src_files[0]) if src_files else None
        tgt_path = str(tgt_files[0]) if tgt_files else None

        if tgt_path or src_path:
            src_cols, tgt_cols = _read_file_headers(src_path or "", tgt_path or "")
            src_set = set(src_cols)
            tgt_set = set(tgt_cols)

            # Target columns are the master list
            # Mark each as: in_target=True always, in_source=True if also in source
            for col in sorted(tgt_set):
                available.append({
                    "field":     col,
                    "label":     get_label(col, custom),
                    "in_source": col in src_set,
                    "in_target": True,
                    "common":    col in src_set,
                    "selected":  not sel_set or col in sel_set,
                })
            # Add source-only columns (in source but not target)
            for col in sorted(src_set - tgt_set):
                available.append({
                    "field":     col,
                    "label":     get_label(col, custom),
                    "in_source": True,
                    "in_target": False,
                    "common":    False,
                    "selected":  False,
                })
    except Exception as e:
        log_event(f"Config: could not read file headers from disk — {e}", "warn")

    # Fallback to last scan result if disk read produced nothing
    if not available and results_store:
        first     = next(iter(results_store.values()))
        available = first.get("available_fields", [
            {"field": fr["field"], "label": get_label(fr["field"], custom),
             "in_source": True, "in_target": True, "common": True,
             "selected":  not sel_set or fr["field"] in sel_set}
            for fr in first.get("field_results", [])
        ])

    # Also include which files are currently on disk so the UI can show them
    src_dir2, tgt_dir2 = get_dirs()
    src_files_list = sorted([f.name for f in src_dir2.iterdir()
                              if f.suffix.lower() in SUPPORTED_EXT])
    tgt_files_list = sorted([f.name for f in tgt_dir2.iterdir()
                              if f.suffix.lower() in SUPPORTED_EXT])

    return jsonify({
        "source_dir":         cfg.get("source_dir",      DEFAULT_CONFIG["source_dir"]),
        "target_dir":         cfg.get("target_dir",      DEFAULT_CONFIG["target_dir"]),
        "pass_threshold":     cfg.get("pass_threshold",  100.0),
        "selected_fields":    sel,
        "active_template":    cfg.get("active_template", ""),
        "available_fields":   available,
        "source_files":       src_files_list,
        "target_files":       tgt_files_list,
        "labels_file_exists": LABELS_FILE.exists(),
        "labels_file":        str(LABELS_FILE) if LABELS_FILE.exists() else None,
    })


@app.route("/api/config", methods=["POST"])
def api_set_config():
    data    = request.get_json(force=True)
    cfg     = load_config()
    changed = False

    for key in ("source_dir", "target_dir"):
        if key in data and str(data[key]).strip():
            np = str(Path(str(data[key]).strip()))
            if np != cfg.get(key):
                cfg[key] = np
                changed  = True

    if "pass_threshold" in data:
        thr = float(data["pass_threshold"])
        if thr != cfg.get("pass_threshold"):
            cfg["pass_threshold"] = thr
            changed = True
            log_event(f"Pass threshold → {thr}%", "info")

    if "selected_fields" in data:
        sel = [str(f).strip().upper() for f in data["selected_fields"]
               if str(f).strip()]
        if sel != cfg.get("selected_fields", []):
            cfg["selected_fields"] = sel
            changed = True
            log_event(
                f"Field selection: {len(sel)} fields" if sel
                else "Field selection: all fields",
                "info",
            )

    if changed:
        save_config(cfg)
        results_store.clear()
        for n in file_states:
            if file_states[n].get("state") == "done":
                file_states[n]["state"] = "changed"
        threading.Thread(target=scan_and_validate_all, daemon=True).start()

    return jsonify({"ok": True, "config": cfg})


# ── Field preview ────────────────────────────────────────────────────────────────

@app.route("/api/fields/from-files", methods=["POST"])
def api_fields_from_files():
    """
    Read headers from specific files already on disk by filename.
    Body: {"source_file": "customers.csv", "target_file": "Export_Data.csv"}
    Called when user picks a specific file pair to load fields from.
    Returns full field list with labels immediately — no scan needed.
    """
    data     = request.get_json(force=True)
    src_name = data.get("source_file", "").strip()
    tgt_name = data.get("target_file", "").strip()
    custom   = _get_custom_labels()
    sel      = load_config().get("selected_fields", [])
    sel_set  = set(sel)

    src_dir, tgt_dir = get_dirs()
    src_path = str(src_dir / src_name) if src_name else None
    tgt_path = str(tgt_dir / tgt_name) if tgt_name else None

    errors = {}
    src_cols, tgt_cols = [], []

    if src_path and Path(src_path).exists():
        try:
            src_cols, _ = _read_file_headers(src_path)
        except Exception as e:
            errors["source"] = str(e)
    elif src_name:
        errors["source"] = f"File not found: {src_name}"

    if tgt_path and Path(tgt_path).exists():
        try:
            _, tgt_cols = _read_file_headers(None, tgt_path)
        except Exception as e:
            errors["target"] = str(e)
    elif tgt_name:
        errors["target"] = f"File not found: {tgt_name}"

    src_set  = set(src_cols)
    tgt_set  = set(tgt_cols)

    fields = []
    # TARGET columns are the master list — show all target fields first
    for col in sorted(tgt_set):
        fields.append({
            "field":     col,
            "label":     get_label(col, custom),
            "in_source": col in src_set,
            "in_target": True,
            "common":    col in src_set,
            "selected":  not sel_set or col in sel_set,
        })
    # Source-only columns (in source but not in target)
    for col in sorted(src_set - tgt_set):
        fields.append({
            "field":     col,
            "label":     get_label(col, custom),
            "in_source": True,
            "in_target": False,
            "common":    False,
            "selected":  False,
        })

    common   = len(src_set & tgt_set)
    src_only = len(src_set - tgt_set)
    tgt_only = len(tgt_set - src_set)

    log_event(
        f"Fields from files: "
        f"src={src_name}({len(src_cols)}) "
        f"tgt={tgt_name}({len(tgt_cols)}) "
        f"→ {common} common, {src_only} src-only, {tgt_only} tgt-only",
        "info",
    )

    return jsonify({
        "fields":      fields,
        "src_count":   len(src_cols),
        "tgt_count":   len(tgt_cols),
        "common":      common,
        "src_only":    src_only,
        "tgt_only":    tgt_only,
        "errors":      errors,
        "source_file": src_name,
        "target_file": tgt_name,
    })


@app.route("/api/fields/preview", methods=["POST"])
def api_fields_preview():
    data      = request.get_json(force=True)
    src_path  = data.get("source_path", "").strip()
    tgt_path  = data.get("target_path", "").strip()
    src_delim = data.get("source_delimiter", ",")
    tgt_delim = data.get("target_delimiter", ",")
    custom    = _get_custom_labels()

    def read_headers(path, delim):
        p = Path(path)
        if not p.exists():
            return None, f"File not found: {path}"
        try:
            if p.suffix.lower() in (".xlsx", ".xls"):
                import openpyxl
                wb   = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                ws   = wb.active
                cols = [str(c.value).strip().upper()
                        for c in next(ws.iter_rows(max_row=1)) if c.value]
                wb.close()
            else:
                import csv
                with open(str(p), encoding="utf-8-sig") as f:
                    cols = [c.strip().upper()
                            for c in next(csv.reader(f, delimiter=delim))]
            return cols, None
        except Exception as e:
            return None, str(e)

    src_cols, src_err = read_headers(src_path, src_delim) if src_path else ([], None)
    tgt_cols, tgt_err = read_headers(tgt_path, tgt_delim) if tgt_path else ([], None)

    errors = {}
    if src_err: errors["source"] = src_err
    if tgt_err: errors["target"] = tgt_err

    src_set  = set(src_cols or [])
    tgt_set  = set(tgt_cols or [])
    common   = sorted(src_set & tgt_set)
    src_only = sorted(src_set - tgt_set)
    tgt_only = sorted(tgt_set - src_set)

    cfg          = load_config()
    selected_set = set(cfg.get("selected_fields", []))

    fields = []
    for col in common:
        fields.append({
            "field": col, "label": get_label(col, custom),
            "in_source": True, "in_target": True, "common": True,
            "selected": not selected_set or col in selected_set,
        })
    for col in src_only:
        fields.append({
            "field": col, "label": get_label(col, custom),
            "in_source": True, "in_target": False, "common": False,
            "selected": False,
        })
    for col in tgt_only:
        fields.append({
            "field": col, "label": get_label(col, custom),
            "in_source": False, "in_target": True, "common": False,
            "selected": False,
        })

    return jsonify({
        "fields":    fields,
        "src_count": len(src_cols or []),
        "tgt_count": len(tgt_cols or []),
        "common":    len(common),
        "src_only":  len(src_only),
        "tgt_only":  len(tgt_only),
        "errors":    errors,
    })


# ── Files & pairs ────────────────────────────────────────────────────────────────

@app.route("/api/files/list")
def api_files_list():
    src, tgt = get_available_files()
    return jsonify({
        "source_files": [f.name for f in src],
        "target_files": [f.name for f in tgt],
    })


@app.route("/api/pairs", methods=["GET"])
def api_pairs_get():
    return jsonify(load_config().get("manual_pairs", []))


@app.route("/api/pairs", methods=["POST"])
def api_pairs_save():
    data  = request.get_json(force=True)
    seen  = set()
    clean = []
    for p in data.get("pairs", []):
        name = str(p.get("name", "")).strip().upper()
        sf   = str(p.get("source_file", "")).strip()
        tf   = str(p.get("target_file", "")).strip()
        if not name or not sf or not tf or name in seen:
            continue
        seen.add(name)
        clean.append({"name": name, "source_file": sf, "target_file": tf})
    cfg = load_config()
    cfg["manual_pairs"] = clean
    save_config(cfg)
    results_store.clear()
    for n in list(file_states):
        if file_states[n].get("state") == "done":
            file_states[n]["state"] = "changed"
    log_event(f"Manual pairs updated: {len(clean)} pair(s)", "info")
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True, "saved": len(clean)})


@app.route("/api/pairs/<name>", methods=["DELETE"])
def api_pairs_delete(name):
    cfg    = load_config()
    pairs  = cfg.get("manual_pairs", [])
    before = len(pairs)
    pairs  = [p for p in pairs if p["name"].upper() != name.upper()]
    cfg["manual_pairs"] = pairs
    save_config(cfg)
    removed = before - len(pairs)
    if removed:
        results_store.pop(name.upper(), None)
        file_states.pop(name.upper(), None)
        log_event(f"Manual pair removed: {name}", "info")
    return jsonify({"ok": True, "removed": removed})


# ── Join key routes ────────────────────────────────────────────────────────────
# These routes power the "Select Join Keys" panel in the dashboard.
# No keys are ever hardcoded — everything comes from the actual uploaded files.

@app.route("/api/join-keys", methods=["GET"])
def api_join_keys_get():
    """Return all saved manual join keys keyed by pair name."""
    return jsonify(load_config().get("manual_join_keys", {}))


@app.route("/api/join-keys/<name>/columns", methods=["GET"])
def api_join_keys_columns(name):
    """
    Return the common columns between source and target files for a pair.
    This is what the UI shows as checkboxes for the user to pick join keys from.
    No hardcoding — purely based on what is in the uploaded files.
    """
    pname = name.upper()
    pairs = discover_pairs()
    pair  = next((p for p in pairs if p["name"] == pname), None)
    if not pair or not pair["has_pair"]:
        return jsonify({"error": f"Pair '{pname}' not found or missing files"}), 404

    def read_headers(path):
        import csv as _csv
        p = str(path)
        if p.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb   = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws   = wb.active
            cols = [str(c.value).strip().upper()
                    for c in next(ws.iter_rows(max_row=1)) if c.value]
            wb.close()
            return cols
        with open(p, encoding="utf-8-sig") as f:
            return [c.strip().upper() for c in next(_csv.reader(f))]

    try:
        src_hdrs = read_headers(pair["source_path"])
        tgt_hdrs = read_headers(pair["target_path"])
    except Exception as e:
        return jsonify({"error": f"Cannot read file headers: {e}"}), 500

    custom      = _get_custom_labels()
    src_set     = set(src_hdrs)
    tgt_set     = set(tgt_hdrs)
    common      = sorted(src_set & tgt_set)
    src_only    = sorted(src_set - tgt_set)
    tgt_only    = sorted(tgt_set - src_set)

    # Current saved keys for this pair
    saved_keys = load_config().get("manual_join_keys", {}).get(pname, [])

    return jsonify({
        "pair_name":     pname,
        "source_file":   pair["source_file"],
        "target_file":   pair["target_file"],
        "common_columns": [
            {
                "field":    col,
                "label":    get_label(col, custom),
                "selected": col in saved_keys,
            }
            for col in common
        ],
        "source_only":   [{"field": c, "label": get_label(c, custom)} for c in src_only],
        "target_only":   [{"field": c, "label": get_label(c, custom)} for c in tgt_only],
        "saved_keys":    saved_keys,
        "total_common":  len(common),
    })


@app.route("/api/join-keys/<name>", methods=["POST"])
def api_join_keys_set(name):
    """
    Save user-selected join keys for a pair.
    Body: {"keys": ["MATNR","KSCHL","EKORG"]}
    Triggers immediate re-validation using the new composite key.
    """
    data  = request.get_json(force=True)
    keys  = [k.strip().upper() for k in data.get("keys", []) if k.strip()]
    cfg   = load_config()
    pname = name.upper()

    if "manual_join_keys" not in cfg:
        cfg["manual_join_keys"] = {}

    if keys:
        cfg["manual_join_keys"][pname] = keys
        log_event(
            f"Join keys set for {pname}: {' + '.join(keys)} "
            f"({len(keys)} field composite key)",
            "info",
        )
    else:
        cfg["manual_join_keys"].pop(pname, None)
        log_event(f"Join keys cleared for {pname} — auto-detect will suggest", "info")

    save_config(cfg)
    results_store.pop(pname, None)
    if pname in file_states:
        file_states[pname]["state"] = "changed"
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True, "name": pname, "keys": keys})


@app.route("/api/join-keys/<name>", methods=["DELETE"])
def api_join_keys_clear(name):
    """Clear saved join keys — auto-detection will suggest on next validation."""
    cfg   = load_config()
    pname = name.upper()
    cfg.get("manual_join_keys", {}).pop(pname, None)
    save_config(cfg)
    results_store.pop(pname, None)
    if pname in file_states:
        file_states[pname]["state"] = "changed"
    log_event(f"Join keys cleared for {pname}", "info")
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/join-keys/<name>/suggest", methods=["POST"])
def api_join_keys_suggest(name):
    """
    Run auto-detection on the uploaded files and return suggested keys.
    Does NOT save anything — purely a suggestion for the UI to show.
    The user still has to click 'Apply' to actually save the keys.
    """
    pname = name.upper()
    pairs = discover_pairs()
    pair  = next((p for p in pairs if p["name"] == pname), None)
    if not pair or not pair["has_pair"]:
        return jsonify({"error": f"Pair '{pname}' not found"}), 404

    try:
        import pandas as pd
        from core.key_detector import detect_composite_key

        def read_headers(path):
            import csv as _csv
            p = str(path)
            if p.lower().endswith((".xlsx", ".xls")):
                import openpyxl
                wb   = openpyxl.load_workbook(p, read_only=True, data_only=True)
                ws   = wb.active
                cols = [str(c.value).strip().upper()
                        for c in next(ws.iter_rows(max_row=1)) if c.value]
                wb.close()
                return cols
            with open(p, encoding="utf-8-sig") as f:
                return [c.strip().upper() for c in next(_csv.reader(f))]

        src_hdrs = read_headers(pair["source_path"])
        tgt_hdrs = read_headers(pair["target_path"])
        common   = sorted(set(src_hdrs) & set(tgt_hdrs))

        # Sample up to 5000 rows — enough for uniqueness scoring without full load
        def load_sample(path, cols):
            p = str(path)
            if p.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(p, dtype=str, nrows=5000,
                                   usecols=cols if cols else None)
            else:
                df = pd.read_csv(p, dtype=str, encoding="utf-8-sig",
                                 nrows=5000, na_filter=False,
                                 usecols=cols if cols else None)
            df.columns = df.columns.str.strip().str.upper()
            return df

        src_df = load_sample(pair["source_path"], common)
        tgt_df = load_sample(pair["target_path"], common)

        kd     = detect_composite_key(src_df, tgt_df, object_name=pname)
        custom = _get_custom_labels()

        # Show uniqueness score for each common column individually
        col_scores = []
        for col in common:
            if col in src_df.columns and col in tgt_df.columns:
                src_unique = src_df[col].nunique() / max(len(src_df), 1)
                tgt_unique = tgt_df[col].nunique() / max(len(tgt_df), 1)
                col_scores.append({
                    "field":          col,
                    "label":          get_label(col, custom),
                    "src_uniqueness": round(src_unique * 100, 1),
                    "tgt_uniqueness": round(tgt_unique * 100, 1),
                    "in_suggestion":  col in kd.join_keys,
                })
        col_scores.sort(key=lambda x: -x["src_uniqueness"])

        return jsonify({
            "ok":               True,
            "suggested_keys":   kd.join_keys,
            "key_labels":       {k: get_label(k, custom) for k in kd.join_keys},
            "detection_method": kd.detection_method,
            "confidence":       kd.confidence,
            "uniqueness_src":   round(kd.uniqueness_src * 100, 1),
            "uniqueness_tgt":   round(kd.uniqueness_tgt * 100, 1),
            "duplicate_src":    kd.duplicate_src,
            "duplicate_tgt":    kd.duplicate_tgt,
            "common_columns":   common,
            "column_scores":    col_scores,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── LTMC XML (SAP Migration Cockpit) routes ───────────────────────────────────

# Store parsed LTMC sheets in memory (keyed by upload session)
_ltmc_store: dict = {}   # {filename: {sheet_name: DataFrame}}


@app.route("/api/ltmc/upload", methods=["POST"])
def api_ltmc_upload():
    """
    Upload a SAP LTMC SpreadsheetML XML file.
    Parses all worksheets immediately and returns a summary of sheets found.
    Each worksheet becomes a separate validation target.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400

    save_name = secure_filename(f.filename)
    if not save_name.lower().endswith(".xml"):
        return jsonify({"error": "Please upload a .xml file (SAP LTMC SpreadsheetML export)"}), 400

    # Save to source directory
    src_dir, _ = get_dirs()
    save_path  = src_dir / save_name
    f.save(str(save_path))
    log_event(f"LTMC XML uploaded: {save_name}", "info")

    try:
        from core.ltmc_parser import parse_ltmc_xml, get_sheet_summary
        sheets  = parse_ltmc_xml(str(save_path))
        summary = get_sheet_summary(sheets)

        # Store parsed sheets
        _ltmc_store[save_name] = sheets

        log_event(
            f"LTMC XML parsed: {save_name} — "
            f"{len(sheets)} sheet(s): {', '.join(sheets.keys())}",
            "info",
        )
        return jsonify({
            "ok":        True,
            "filename":  save_name,
            "sheets":    summary,
            "total_sheets": len(sheets),
        })
    except Exception as e:
        log_event(f"LTMC parse error ({save_name}): {e}", "error")
        return jsonify({"error": f"Failed to parse XML: {e}"}), 500


@app.route("/api/ltmc/sheets/<filename>", methods=["GET"])
def api_ltmc_sheets(filename):
    """
    Return the list of sheets and their columns for a parsed LTMC file.
    Used to let the user pick which sheet to validate against which post-load file.
    """
    safe = secure_filename(filename)
    src_dir, _ = get_dirs()
    xml_path   = src_dir / safe

    # Re-parse if not in memory
    if safe not in _ltmc_store:
        if not xml_path.exists():
            return jsonify({"error": f"File not found: {safe}"}), 404
        try:
            from core.ltmc_parser import parse_ltmc_xml
            _ltmc_store[safe] = parse_ltmc_xml(str(xml_path))
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    sheets  = _ltmc_store[safe]
    custom  = _get_custom_labels()
    result  = []
    for sheet_name, df in sheets.items():
        result.append({
            "sheet_name":   sheet_name,
            "row_count":    len(df),
            "columns":      list(df.columns),
            "col_count":    len(df.columns),
            "col_labels":   {c: get_label(c, custom) for c in df.columns},
        })
    return jsonify({"filename": safe, "sheets": result})


@app.route("/api/ltmc/resolve", methods=["POST"])
def api_ltmc_resolve():
    """
    Given a LTMC sheet's columns and a post-load file's columns,
    resolve the post-load friendly names to SAP technical names and
    show which columns matched and which didn't.

    Body:
    {
      "ltmc_file":    "migration_template.xml",
      "sheet_name":   "CustomerHeader",
      "postload_file":"customer_extract.csv"   // already in target dir
    }
    """
    data        = request.get_json(force=True)
    ltmc_file   = data.get("ltmc_file", "").strip()
    sheet_name  = data.get("sheet_name", "").strip()
    postload_file = data.get("postload_file", "").strip()

    if not ltmc_file or not sheet_name or not postload_file:
        return jsonify({"error": "ltmc_file, sheet_name, and postload_file are required"}), 400

    safe_ltmc = secure_filename(ltmc_file)
    safe_post = secure_filename(postload_file)
    src_dir, tgt_dir = get_dirs()

    # Get LTMC sheet
    if safe_ltmc not in _ltmc_store:
        xml_path = src_dir / safe_ltmc
        if not xml_path.exists():
            return jsonify({"error": f"LTMC file not found: {safe_ltmc}"}), 404
        try:
            from core.ltmc_parser import parse_ltmc_xml
            _ltmc_store[safe_ltmc] = parse_ltmc_xml(str(xml_path))
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    sheets = _ltmc_store[safe_ltmc]
    if sheet_name not in sheets:
        # Try case-insensitive match
        matched = next((k for k in sheets if k.lower() == sheet_name.lower()), None)
        if not matched:
            return jsonify({
                "error": f"Sheet '{sheet_name}' not found in {safe_ltmc}. "
                         f"Available: {list(sheets.keys())}"
            }), 404
        sheet_name = matched

    ltmc_df = sheets[sheet_name]

    # Load post-load file headers
    post_path = tgt_dir / safe_post
    if not post_path.exists():
        return jsonify({"error": f"Post-load file not found: {safe_post}"}), 404

    try:
        if str(post_path).lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb      = openpyxl.load_workbook(str(post_path), read_only=True, data_only=True)
            ws_post = wb.active
            post_cols = [str(c.value).strip() for c in next(ws_post.iter_rows(max_row=1)) if c.value]
            wb.close()
        else:
            import csv as csv_mod
            with open(str(post_path), encoding="utf-8-sig") as fh:
                post_cols = [c.strip() for c in next(csv_mod.reader(fh))]
    except Exception as e:
        return jsonify({"error": f"Cannot read post-load file headers: {e}"}), 500

    # Resolve post-load columns to SAP technical names
    from core.label_resolver import resolve_postload_columns
    from core.field_labels import SAP_FIELD_LABELS
    custom = _get_custom_labels()

    resolution = resolve_postload_columns(
        postload_columns=post_cols,
        ltmc_columns=list(ltmc_df.columns),
        field_labels=SAP_FIELD_LABELS,
        custom_labels=custom,
    )

    matched     = {col: info for col, info in resolution.items() if info["matched"]}
    unmatched   = {col: info for col, info in resolution.items() if not info["matched"]}
    ltmc_unmatched = [c for c in ltmc_df.columns
                      if c not in {i["ltmc_col"] for i in matched.values()}]

    log_event(
        f"LTMC resolve: {sheet_name} vs {safe_post} — "
        f"{len(matched)} matched, {len(unmatched)} unmatched postload cols, "
        f"{len(ltmc_unmatched)} unmatched LTMC cols",
        "info",
    )

    return jsonify({
        "ok":                  True,
        "ltmc_file":           safe_ltmc,
        "sheet_name":          sheet_name,
        "postload_file":       safe_post,
        "ltmc_row_count":      len(ltmc_df),
        "ltmc_columns":        list(ltmc_df.columns),
        "postload_columns":    post_cols,
        "matched":             {k: v for k, v in matched.items()},
        "unmatched_postload":  {k: v for k, v in unmatched.items()},
        "unmatched_ltmc":      ltmc_unmatched,
        "match_count":         len(matched),
        "coverage_pct":        round(len(matched) / max(len(ltmc_df.columns), 1) * 100, 1),
    })


@app.route("/api/ltmc/validate", methods=["POST"])
def api_ltmc_validate():
    """
    Run full validation: LTMC sheet (source) vs post-load file (target).

    Body:
    {
      "ltmc_file":        "migration_template.xml",
      "sheet_name":       "CustomerHeader",
      "postload_file":    "customer_extract.csv",
      "join_keys":        ["KUNNR"],           // optional — auto-detected if not provided
      "pass_threshold":   90.0                 // optional
    }

    Flow:
    1. Parse LTMC sheet → source DataFrame (SAP technical names)
    2. Load post-load file → target DataFrame (friendly or technical names)
    3. Resolve post-load column names to SAP technical names
    4. Build field_map {ltmc_col: postload_col}
    5. Save LTMC sheet as temp CSV so MaterialValidator can use it
    6. Run MaterialValidator with composite key
    7. Return full validation result
    """
    import tempfile
    data          = request.get_json(force=True)
    ltmc_file     = data.get("ltmc_file", "").strip()
    sheet_name    = data.get("sheet_name", "").strip()
    postload_file = data.get("postload_file", "").strip()
    join_keys     = [k.strip().upper() for k in data.get("join_keys", []) if k.strip()]
    pass_threshold= float(data.get("pass_threshold",
                          load_config().get("pass_threshold", 100.0)))

    if not ltmc_file or not sheet_name or not postload_file:
        return jsonify({"error": "ltmc_file, sheet_name, postload_file required"}), 400

    safe_ltmc = secure_filename(ltmc_file)
    safe_post = secure_filename(postload_file)
    src_dir, tgt_dir = get_dirs()

    # ── Get LTMC sheet ──────────────────────────────────────────────────────
    if safe_ltmc not in _ltmc_store:
        xml_path = src_dir / safe_ltmc
        if not xml_path.exists():
            return jsonify({"error": f"LTMC file not found: {safe_ltmc}"}), 404
        try:
            from core.ltmc_parser import parse_ltmc_xml
            _ltmc_store[safe_ltmc] = parse_ltmc_xml(str(xml_path))
        except Exception as e:
            return jsonify({"error": f"XML parse error: {e}"}), 500

    sheets = _ltmc_store[safe_ltmc]
    if sheet_name not in sheets:
        matched_name = next((k for k in sheets if k.lower() == sheet_name.lower()), None)
        if not matched_name:
            return jsonify({"error": f"Sheet '{sheet_name}' not found"}), 404
        sheet_name = matched_name

    ltmc_df = sheets[sheet_name].copy()

    # ── Load post-load file ─────────────────────────────────────────────────
    post_path = tgt_dir / safe_post
    if not post_path.exists():
        return jsonify({"error": f"Post-load file not found: {safe_post}"}), 404

    try:
        if str(post_path).lower().endswith((".xlsx", ".xls")):
            post_df = pd.read_excel(str(post_path), dtype=str)
        else:
            post_df = pd.read_csv(str(post_path), dtype=str,
                                  encoding="utf-8-sig", na_filter=False)
        for col in post_df.select_dtypes(include="object").columns:
            post_df[col] = post_df[col].str.strip()
    except Exception as e:
        return jsonify({"error": f"Cannot load post-load file: {e}"}), 500

    # ── Resolve post-load columns to SAP technical names ───────────────────
    from core.label_resolver import (
        resolve_postload_columns, build_field_map_from_resolution
    )
    from core.field_labels import SAP_FIELD_LABELS
    custom = _get_custom_labels()

    resolution = resolve_postload_columns(
        postload_columns=list(post_df.columns),
        ltmc_columns=list(ltmc_df.columns),
        field_labels=SAP_FIELD_LABELS,
        custom_labels=custom,
    )

    # Rename post-load columns to their resolved SAP names for validation
    rename_map = {}
    for orig_col, info in resolution.items():
        if info["matched"] and info["resolved"] != orig_col:
            rename_map[orig_col] = info["resolved"]

    post_df_renamed = post_df.rename(columns=rename_map)

    # Build field_map: {ltmc_col (=SAP code): postload_resolved_col (=SAP code)}
    # After renaming, both sides use SAP technical names → simple exact map
    ltmc_cols  = set(ltmc_df.columns)
    post_cols  = set(post_df_renamed.columns)
    field_map  = {c: c for c in ltmc_cols & post_cols}

    if not field_map:
        return jsonify({
            "error": "No matching columns found between LTMC sheet and post-load file. "
                     "Check that the post-load file contains the same SAP fields "
                     "as the LTMC template (even with different names).",
            "ltmc_columns":     list(ltmc_df.columns),
            "postload_columns": list(post_df.columns),
            "resolution":       {k: v for k, v in list(resolution.items())[:20]},
        }), 422

    log_event(
        f"LTMC validate: {sheet_name} ({len(ltmc_df)} rows) vs "
        f"{safe_post} ({len(post_df)} rows) — "
        f"{len(field_map)} fields mapped",
        "info",
    )

    # ── Save DataFrames as temp CSVs for MaterialValidator ─────────────────
    import tempfile, os
    tmp_dir = Path(tempfile.mkdtemp())
    src_csv = tmp_dir / f"ltmc_{sheet_name}.csv"
    tgt_csv = tmp_dir / f"postload_{safe_post}.csv"

    ltmc_df.to_csv(str(src_csv), index=False)
    post_df_renamed.to_csv(str(tgt_csv), index=False)

    # ── Run MaterialValidator ───────────────────────────────────────────────
    from core.validator import MaterialValidator
    validator = MaterialValidator(
        field_map=field_map,
        manual_join_keys=join_keys if join_keys else None,
        pass_threshold=pass_threshold,
        custom_labels=custom if custom else None,
    )

    try:
        result = validator.validate(
            str(src_csv), str(tgt_csv),
            object_name=sheet_name,
        )
    except Exception as e:
        return jsonify({"error": f"Validation failed: {e}"}), 500
    finally:
        # Cleanup temp files
        try:
            src_csv.unlink(); tgt_csv.unlink(); tmp_dir.rmdir()
        except Exception:
            pass

    # ── Build response ──────────────────────────────────────────────────────
    ss = result.summary_stats
    business_status = calculate_business_status(result, pass_threshold)

    field_rows = []
    for fr in result.field_results:
        disp = get_display(fr.field_source, fr.field_target, custom)
        # Find original post-load column name for display
        original_postload_col = next(
            (orig for orig, info in resolution.items()
             if info.get("resolved") == fr.field_source and info["matched"]),
            fr.field_source
        )
        field_rows.append({
            "field":                fr.field_source,
            "field_label":          disp["source_label"],
            "field_target":         fr.field_target,
            "postload_original_col":original_postload_col,
            "resolution_method":    resolution.get(original_postload_col, {}).get("method", "exact"),
            "is_key_field":         fr.is_key_field,
            "type":                 "numeric" if fr.is_numeric else "string",
            "tolerance":            fr.tolerance_used,
            "total":                fr.total_records,
            "matched":              fr.matched,
            "mismatched":           fr.mismatched,
            "miss_source":          fr.missing_in_source,
            "miss_target":          fr.missing_in_target,
            "match_pct":            fr.match_pct,
            "pass_threshold":       fr.pass_threshold,
            "status":               fr.status,
            "mismatches":           fr.mismatch_details,
            "matches":              fr.matched_details,
            "mismatch_count":       len(fr.mismatch_details),
        })

    # Save Excel report
    ts             = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"LTMC_{sheet_name}_{ts}.xlsx"
    excel_path     = REPORTS_DIR / excel_filename
    result_dict    = {
        "name":                   f"LTMC_{sheet_name}",
        "ltmc_file":              safe_ltmc,
        "sheet_name":             sheet_name,
        "postload_file":          safe_post,
        "status":                 business_status["status"],
        "business_message":       business_status["message"],
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "pass_threshold":         pass_threshold,
        "join_keys":              result.join_keys,
        "key_detection_method":   result.key_detection_method,
        "key_confidence":         result.key_confidence,
        "duplicate_src":          result.duplicate_src,
        "duplicate_tgt":          result.duplicate_tgt,
        "field_results":          field_rows,
        "resolution_summary": {
            "matched_count":   len([i for i in resolution.values() if i["matched"]]),
            "unmatched_count": len([i for i in resolution.values() if not i["matched"]]),
            "resolution":      {k: v for k, v in resolution.items()},
        },
        "run_at":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "excel_file":  excel_filename,
    }

    try:
        generate_excel_report(result_dict, str(excel_path))
    except Exception as e:
        result_dict["excel_error"] = str(e)

    cache_and_trim_row_details(result_dict)

    log_event(
        f"LTMC validation complete: {sheet_name} — "
        f"{business_status['status']} | "
        f"matched={result.records_matched} "
        f"src_only={result.records_only_in_source} "
        f"tgt_only={result.records_only_in_target}",
        "success" if business_status["status"] == "PASS" else "warn",
    )

    return jsonify(result_dict)


@app.route("/api/ltmc/list", methods=["GET"])
def api_ltmc_list():
    """List all uploaded LTMC XML files in the source directory."""
    src_dir, _ = get_dirs()
    xml_files  = sorted(src_dir.glob("*.xml"), key=lambda f: f.stat().st_mtime, reverse=True)
    result     = []
    for f in xml_files:
        sheets = list(_ltmc_store[f.name].keys()) if f.name in _ltmc_store else []
        result.append({
            "filename":   f.name,
            "size_kb":    round(f.stat().st_size / 1024, 1),
            "modified":   datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            "sheets":     sheets,
            "parsed":     f.name in _ltmc_store,
        })
    return jsonify(result)


# ── Labels / reports / downloads ─────────────────────────────────────────────────

@app.route("/api/recommendations/<name>")
def api_recommendations(name):
    result = results_store.get(name.upper())
    if not result:
        return jsonify({"error": "Validation result not found"}), 404
    recs = build_recommendations(result)
    result["recommendations"] = recs
    return jsonify({"recommendations": recs})


@app.route("/api/corrections/<name>/apply", methods=["POST"])
def api_apply_correction(name):
    """Create a corrected target copy after an explicit user-approved field action."""
    result = results_store.get(name.upper())
    if not result:
        return jsonify({"error": "Validation result not found"}), 404
    data = request.get_json(silent=True) or {}
    source_field = str(data.get("field", "")).strip().upper()
    rec = next((r for r in build_recommendations(result)
                if r["field"] == source_field and r["can_apply"]), None)
    if not rec:
        return jsonify({"error": "No safe correction is available for this field"}), 400
    if result.get("duplicate_src", 0) or result.get("duplicate_tgt", 0):
        return jsonify({"error": "Resolve duplicate join keys before applying corrections"}), 409

    pair = next((p for p in discover_pairs()
                 if p.get("has_pair") and p.get("name", "").upper() == name.upper()), None)
    if not pair:
        return jsonify({"error": "Source/target pair not found"}), 404
    join_keys = [str(k).upper() for k in (result.get("join_keys") or [])]
    target_field = rec["target_field"].upper()
    if not join_keys:
        return jsonify({"error": "No validated join key is available"}), 400

    src = _normalise_columns(_load_correction_file(pair["source_path"]))
    tgt = _normalise_columns(_load_correction_file(pair["target_path"]))
    required_src = join_keys + [source_field]
    required_tgt = join_keys + [target_field]
    missing = [c for c in required_src if c not in src.columns] + [c for c in required_tgt if c not in tgt.columns]
    if missing:
        return jsonify({"error": "Required columns missing: " + ", ".join(sorted(set(missing)))}), 400

    def key_series(df):
        return df[join_keys].fillna("").astype(str).apply(
            lambda row: "||".join(v.strip().upper() for v in row), axis=1)

    src_keys = key_series(src)
    tgt_keys = key_series(tgt)
    if src_keys.duplicated().any() or tgt_keys.duplicated().any():
        return jsonify({"error": "Correction stopped because the selected keys are not unique"}), 409
    source_values = dict(zip(src_keys, src[source_field]))
    replacement = tgt_keys.map(source_values)
    matched = replacement.notna()
    before = tgt[target_field].fillna("").astype(str)
    changed = int((matched & (before != replacement.fillna("").astype(str))).sum())
    tgt.loc[matched, target_field] = replacement[matched]

    original = Path(pair["target_path"])
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = REPORTS_DIR / f"{original.stem}_corrected_{stamp}{original.suffix.lower()}"
    if output.suffix in (".xlsx", ".xls"):
        if output.suffix == ".xls":
            output = output.with_suffix(".xlsx")
        tgt.to_excel(output, index=False)
    else:
        tgt.to_csv(output, index=False, encoding="utf-8-sig")
    save_learned_rule(result["name"], source_field, target_field)
    correction_rag.remember(
        object_name=result["name"], source_field=source_field, target_field=target_field,
        issues=[e.get("issue", "") for e in rec.get("examples", [])],
        examples=rec.get("examples", []), action="copy_source_to_target",
    )
    result["recommendations"] = build_recommendations(result)
    corrected_files = result.setdefault("corrected_files", [])
    corrected_files.insert(0, {
        "filename": output.name,
        "download_url": "/api/download-corrected/" + output.name,
        "changed_records": changed,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    del corrected_files[10:]
    log_event(f"{name}: created corrected target copy for {target_field}; {changed} values changed", "success")
    return jsonify({
        "ok": True, "changed_records": changed, "filename": output.name,
        "download_url": "/api/download-corrected/" + output.name,
        "message": "Correction learned and applied to a new target copy. Original upload was not changed.",
    })


@app.route("/api/download-corrected/<filename>")
def api_download_corrected(filename):
    safe = secure_filename(filename)
    path = REPORTS_DIR / safe
    if safe != filename or not path.exists() or path.suffix.lower() not in SUPPORTED_EXT:
        return jsonify({"error": "Corrected file not found"}), 404
    return send_file(str(path), as_attachment=True, download_name=path.name)


@app.route("/api/corrections/<name>/auto-apply", methods=["POST"])
def api_auto_apply_learned(name):
    """Apply only high-confidence retrieved, previously approved rules to one new copy."""
    result = results_store.get(name.upper())
    if not result:
        return jsonify({"error": "Validation result not found"}), 404
    recs = [r for r in build_recommendations(result)
            if r.get("can_apply") and r.get("rag_match") and r.get("rag_confidence", 0) >= 85]
    if not recs:
        return jsonify({"error": "No high-confidence approved rules match this validation"}), 400
    if result.get("duplicate_src", 0) or result.get("duplicate_tgt", 0):
        return jsonify({"error": "Resolve duplicate join keys before applying learned rules"}), 409
    pair = next((p for p in discover_pairs()
                 if p.get("has_pair") and p.get("name", "").upper() == name.upper()), None)
    join_keys = [str(k).upper() for k in (result.get("join_keys") or [])]
    if not pair or not join_keys:
        return jsonify({"error": "Validated source/target pair and join keys are required"}), 400

    src = _normalise_columns(_load_correction_file(pair["source_path"]))
    tgt = _normalise_columns(_load_correction_file(pair["target_path"]))

    def key_series(df):
        return df[join_keys].fillna("").astype(str).apply(
            lambda row: "||".join(v.strip().upper() for v in row), axis=1)
    if any(k not in src.columns or k not in tgt.columns for k in join_keys):
        return jsonify({"error": "Join-key columns are missing"}), 400
    src_keys, tgt_keys = key_series(src), key_series(tgt)
    if src_keys.duplicated().any() or tgt_keys.duplicated().any():
        return jsonify({"error": "Learned-rule correction stopped because keys are not unique"}), 409

    changes = []
    for rec in recs:
        sf, tf = rec["field"].upper(), rec["target_field"].upper()
        if sf not in src.columns or tf not in tgt.columns:
            continue
        replacement = tgt_keys.map(dict(zip(src_keys, src[sf])))
        matched = replacement.notna()
        before = tgt[tf].fillna("").astype(str)
        changed = int((matched & (before != replacement.fillna("").astype(str))).sum())
        if changed:
            tgt.loc[matched, tf] = replacement[matched]
            changes.append({"field": tf, "changed_records": changed,
                            "rag_confidence": rec["rag_confidence"]})
    if not changes:
        return jsonify({"error": "Retrieved rules found no values requiring correction"}), 400

    original = Path(pair["target_path"])
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = REPORTS_DIR / f"{original.stem}_rag_corrected_{stamp}{original.suffix.lower()}"
    if output.suffix in (".xlsx", ".xls"):
        if output.suffix == ".xls": output = output.with_suffix(".xlsx")
        tgt.to_excel(output, index=False)
    else:
        tgt.to_csv(output, index=False, encoding="utf-8-sig")
    result.setdefault("corrected_files", []).insert(0, {
        "filename": output.name, "download_url": "/api/download-corrected/" + output.name,
        "changed_records": sum(c["changed_records"] for c in changes),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    log_event(f"{name}: RAG applied {len(changes)} approved rules to a new target copy", "success")
    return jsonify({
        "ok": True, "filename": output.name,
        "download_url": "/api/download-corrected/" + output.name,
        "changes": changes,
        "message": f"Applied {len(changes)} retrieved approved rule(s). Original target was not changed.",
    })

@app.route("/api/labels/sample")
def api_labels_sample():
    lines = ["FIELD_NAME,FRIENDLY_LABEL"] + [
        f"{k},{v}" for k, v in list(SAP_FIELD_LABELS.items())[:25]
    ]
    return Response(
        "\n".join(lines).encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=sample_labels.csv"},
    )


@app.route("/api/objects")
def api_objects():
    custom = _get_custom_labels()
    return jsonify([
        {
            "key":            k,
            "description":    v.get("description", k),
            "join_key":       v.get("join_key", ""),
            "join_key_label": get_label(v.get("join_key", ""), custom),
            "key_fields":     [
                {"field": f, "label": get_label(f, custom)}
                for f in v.get("key_fields", [])
            ],
        }
        for k, v in SAP_OBJECT_CONFIG.items()
    ])


@app.route("/api/download/<name>")
def api_download(name):
    r = results_store.get(name.upper())
    if not r:
        return jsonify({"error": "Not found"}), 404
    path = REPORTS_DIR / r.get("excel_file", "")
    if not path.exists():
        return jsonify({"error": "Report file missing"}), 404
    return send_file(
        str(path), as_attachment=True, download_name=path.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/download-file/<filename>")
def api_download_file(filename):
    path = REPORTS_DIR / filename
    if not path.exists() or not filename.endswith(".xlsx"):
        return jsonify({"error": "Not found"}), 404
    return send_file(
        str(path), as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/reports")
def api_reports():
    files = sorted(
        REPORTS_DIR.glob("*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    return jsonify([
        {
            "filename": f.name,
            "size_kb":  round(f.stat().st_size / 1024, 1),
            "modified": datetime.fromtimestamp(
                f.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        }
        for f in files
    ])


@app.route("/api/folders")
def api_folders():
    s, t = get_dirs()
    return jsonify({
        "source_dir":  str(s),
        "target_dir":  str(t),
        "reports_dir": str(REPORTS_DIR),
    })


@app.route("/api/clear-results", methods=["POST"])
def api_clear_results():
    results_store.clear()
    file_states.clear()
    activity_log.clear()
    scan_status.update({
        "last_scan": None, "scanning": False, "error": None,
        "current_file": None, "total_files": 0, "completed_files": 0,
    })
    log_event("Results cleared", "info")
    return jsonify({"ok": True})


# ── Entry point ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    s, t = get_dirs()
    cfg  = load_config()
    print("\n  ╔══════════════════════════════════════╗")
    print("  ║  Genpact SAP Migration Validator V4  ║")
    print("  ╚══════════════════════════════════════╝")
    print(f"  Source dir     → {s}")
    print(f"  Target dir     → {t}")
    print(f"  Templates      → {TEMPLATES_DIR}")
    print(f"  Reports        → {REPORTS_DIR}")
    print(f"  Pass threshold → {cfg.get('pass_threshold', 100)}%")
    if cfg.get("active_template"):
        print(f"  Active template→ {cfg['active_template']}")
    port = int(os.environ.get("SAP_VALIDATOR_PORT", "5050"))
    print(f"  Open           → http://127.0.0.1:{port}\n")
    threading.Thread(target=scan_and_validate_all, daemon=True).start()
    # threading.Thread(target=background_watcher, args=(60,), daemon=True).start()
    app.run(debug=False, host="127.0.0.1", port=port, use_reloader=False)
