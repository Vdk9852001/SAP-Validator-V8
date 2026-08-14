"""
Genpact SAP Migration Validator
================================
Desktop tool to validate SAP post-load Excel extracts against XML migration templates.

Features:
  • Parses SAP SpreadsheetML XML (LTMC format)
  • Reads any post-load .xlsx file
  • Auto-maps column headers → SAP field codes
  • Composite join-key detection & editing
  • Field-level diff with status (PASS / WARN / FAIL)
  • Smart search & learn for unmatched keys
  • AI column mapping via Claude API (optional)
  • Full Excel report export (4 sheets)

Requirements:
    pip install openpyxl lxml anthropic

Run:
    python sap_validator.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import threading
import re
import json
import os
import datetime
from xml.etree import ElementTree as ET

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import anthropic
    ANTHROPIC_OK = True
except ImportError:
    ANTHROPIC_OK = False

# ─── Colour palette (matches dashboard) ───────────────────────────────────────
C = {
    "bg":       "#f4f6fa",
    "surface":  "#ffffff",
    "surface2": "#f0f2f7",
    "border":   "#dde1ec",
    "text":     "#1a1f36",
    "muted":    "#6b728e",
    "accent":   "#4f46e5",
    "accent_l": "#eef2ff",
    "pass_":    "#16a34a",
    "pass_bg":  "#dcfce7",
    "fail":     "#dc2626",
    "fail_bg":  "#fee2e2",
    "warn":     "#d97706",
    "warn_bg":  "#fef3c7",
    "info":     "#2563eb",
    "info_bg":  "#dbeafe",
    "teal":     "#0f766e",
    "teal_bg":  "#ccfbf1",
    "header":   "#4f46e5",
    "header_fg":"#ffffff",
}

# ─── SAP field → friendly label dictionary ────────────────────────────────────
SAP_LABELS = {
    "ARBPL":"Work Center","WERKS":"Plant","KTEXT":"Description","VERAN":"Person Responsible",
    "VERWE":"Work Center Category","PLANV":"Usage","STEUS":"Control Key","KTSCH":"Standard Value Key",
    "KOSTL":"Cost Center","AUFNR":"Order","MATNR":"Material","KUNNR":"Customer","LIFNR":"Vendor",
    "BUKRS":"Company Code","VKORG":"Sales Org","VTWEG":"Distribution Channel","SPART":"Division",
    "EKORG":"Purchasing Org","LGORT":"Storage Location","CHARG":"Batch","MEINS":"Base Unit",
    "MATKL":"Material Group","MTART":"Material Type","PRCTR":"Profit Center",
    "LSTAR":"Activity Type","BEGDA":"Valid From","ENDDA":"Valid To",
    "CANUM":"Capacity","KAPAR":"Capacity Category","SPRAS":"Language",
    "PRVBE":"Supply Area","PLANR":"Planner Group","NGRAD":"Utilization %",
    "BEGZT":"Start Time","ENDZT":"End Time","PAUSE":"Break Duration",
    "AZNOR":"Normal Capacity","RGEKZ":"Backflush","ERNAM":"Created By",
    "AENAM":"Changed By","ERSDA":"Created On","LAEDA":"Last Changed",
    "NAME1":"Name 1","NAME2":"Name 2","STRAS":"Street","ORT01":"City",
    "PSTLZ":"Postal Code","LAND1":"Country","REGIO":"Region",
}

# ─── XML Parser ───────────────────────────────────────────────────────────────
NS = "urn:schemas-microsoft-com:office:spreadsheet"

def _tag(name): return f"{{{NS}}}{name}"

def parse_xml_row(row_el, num_cols=120):
    result = [""] * num_cols
    col = 0
    for cell in row_el:
        if cell.tag != _tag("Cell"): continue
        idx = cell.get(_tag("Index"))
        if idx: col = int(idx) - 1
        data = cell.find(_tag("Data"))
        if data is not None and col < num_cols:
            result[col] = (data.text or "").strip()
        col += 1
    return result

def parse_xml_workbook(path):
    """Returns dict: sheet_name → {fields: [...], data: [...]}"""
    tree = ET.parse(path)
    root = tree.getroot()
    sheets = {}
    for ws in root.iter(_tag("Worksheet")):
        name = ws.get(_tag("Name"), "")
        table = ws.find(_tag("Table"))
        if table is None: continue
        rows = [r for r in table if r.tag == _tag("Row")]
        if len(rows) < 4: continue

        # Find SAP field code row (ARBPL, WERKS, etc.)
        field_row_idx, fields = -1, []
        for i, row in enumerate(rows[:10]):
            vals = [v for v in parse_xml_row(row) if v]
            if (len(vals) >= 2
                    and re.match(r'^[A-Z][A-Z0-9_]{2,19}$', vals[0])
                    and sum(1 for v in vals if re.match(r'^[A-Z][A-Z0-9_]{1,19}$', v)) > len(vals) * 0.5):
                field_row_idx = i
                fields = parse_xml_row(row)
                break
        if field_row_idx == -1:
            sheets[name] = {"fields": [], "data": []}
            continue

        # Skip description rows
        data_start = field_row_idx + 1
        while data_start < len(rows):
            first = parse_xml_row(rows[data_start])[0] or ""
            if len(first) > 60 or "\n" in first: data_start += 1
            else: break

        data = []
        for row in rows[data_start:]:
            vals = parse_xml_row(row)
            if not vals[0]: continue
            rec = {fields[i]: vals[i] for i in range(min(len(fields), len(vals))) if fields[i]}
            data.append(rec)

        sheets[name] = {"fields": [f for f in fields if f], "data": data}
    return sheets

# ─── Excel Parser ─────────────────────────────────────────────────────────────
def parse_excel_file(path):
    """Returns {headers: [...], data: [...]}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {"headers": [], "data": []}
    headers = [str(h or "").strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        rec = {headers[i]: str(row[i] if row[i] is not None else "").strip()
               for i in range(min(len(headers), len(row)))}
        if any(rec.values()): data.append(rec)
    return {"headers": headers, "data": data}

# ─── Dynamic Field Mapper ──────────────────────────────────────────────────────
def build_field_map(excel_headers, xml_fields):
    """Auto-maps Excel column headers → XML SAP field codes."""
    label_to_sap = {}
    for code, label in SAP_LABELS.items():
        for key in [label.lower(), label.lower().replace(" ",""), re.sub(r'[^a-z0-9]','',label.lower())]:
            label_to_sap[key] = code

    def norm(s): return re.sub(r'[^a-z0-9]','', (s or "").lower())

    result = {}
    for h in excel_headers:
        hn = norm(h)
        if h in xml_fields:                         result[h] = h; continue
        direct = next((f for f in xml_fields if norm(f) == hn), None)
        if direct:                                  result[h] = direct; continue
        sap = label_to_sap.get(hn) or label_to_sap.get(h.lower())
        if sap and sap in xml_fields:               result[h] = sap; continue
        words = [w for w in hn.split() if len(w) > 3]
        partial = next((f for f in xml_fields
                        if words and all(w in norm(SAP_LABELS.get(f, f)) for w in words)), None)
        if partial:                                 result[h] = partial; continue
        result[h] = None
    return result

# ─── Join Key Detector ─────────────────────────────────────────────────────────
KEY_PATTERNS = [
    ["ARBPL","WERKS"],["MATNR","WERKS"],["KUNNR","VKORG"],["LIFNR","EKORG"],
    ["AUFNR"],["MATNR"],["KUNNR"],["LIFNR"],["ARBPL"],
    ["BUKRS","SAKNR"],["PRCTR","KOKRS"],["KOSTL","KOKRS"],
]

def detect_join_keys(xml_fields):
    for pattern in KEY_PATTERNS:
        if all(k in xml_fields for k in pattern): return list(pattern)
    candidates = [f for f in xml_fields if len(f) <= 6]
    return candidates[:2] if candidates else [xml_fields[0]] if xml_fields else []

# ─── Validation Engine ─────────────────────────────────────────────────────────
def run_validation(xml_sheets, excel_data, excel_headers, join_keys, field_map):
    # Primary sheet
    sheet_name = next((k for k, v in xml_sheets.items() if v["data"]), None)
    if not sheet_name:
        return {"error": "No data found in XML", "records": [], "summary": {}}

    xml_data   = xml_sheets[sheet_name]["data"]
    xml_fields = xml_sheets[sheet_name]["fields"]

    # Map join keys to Excel columns
    excel_join_cols = []
    for xml_key in join_keys:
        excel_col = next((h for h in excel_headers if field_map.get(h) == xml_key), xml_key)
        excel_join_cols.append({"xml_key": xml_key, "excel_col": excel_col})

    # Build XML lookup
    xml_map = {}
    for rec in xml_data:
        key = "||".join((rec.get(k) or "").upper().strip() for k in join_keys)
        xml_map[key] = rec

    # Field pairs to validate (excluding join key fields)
    field_pairs = [
        {"excel_col": h, "xml_field": field_map[h]}
        for h in excel_headers
        if field_map.get(h) and field_map[h] in xml_fields and field_map[h] not in join_keys
    ]

    records = []
    excel_keys = set()
    total_field_diffs = 0

    for excel_rec in excel_data:
        key = "||".join((excel_rec.get(e["excel_col"]) or "").upper().strip()
                        for e in excel_join_cols)
        excel_keys.add(key)
        xml_rec = xml_map.get(key)

        field_checks, has_diff = [], False
        if xml_rec:
            for p in field_pairs:
                ev = (excel_rec.get(p["excel_col"]) or "").strip()
                xv = (xml_rec.get(p["xml_field"]) or "").strip()
                if not ev and not xv: continue
                match = ev.lower() == xv.lower()
                field_checks.append({**p, "excel_val": ev, "xml_val": xv, "match": match})
                if not match: has_diff = True; total_field_diffs += 1

        key_display = " · ".join(excel_rec.get(e["excel_col"]) or "" for e in excel_join_cols)
        records.append({
            "key": key,
            "key_display": key_display,
            "excel_rec": excel_rec,
            "xml_rec": xml_rec,
            "status": "missing_xml" if not xml_rec else ("mismatch" if has_diff else "matched"),
            "field_checks": field_checks,
        })

    xml_only = [r for r in xml_data
                if "||".join((r.get(k) or "").upper().strip() for k in join_keys) not in excel_keys]

    matched    = sum(1 for r in records if r["status"] == "matched")
    mismatch   = sum(1 for r in records if r["status"] == "mismatch")
    missing_xml = sum(1 for r in records if r["status"] == "missing_xml")
    total = len(records)

    return {
        "sheet_name": sheet_name,
        "xml_fields": xml_fields,
        "field_pairs": field_pairs,
        "excel_join_cols": excel_join_cols,
        "records": records,
        "xml_only": xml_only,
        "summary": {
            "total": total,
            "matched": matched,
            "mismatch": mismatch,
            "missing_xml": missing_xml,
            "xml_only": len(xml_only),
            "field_diffs": total_field_diffs,
            "match_rate": round(matched / total * 100) if total else 0,
        },
        "field_map": field_map,
        "join_keys": join_keys,
    }

# ─── Fuzzy Search ─────────────────────────────────────────────────────────────
def fuzzy_search(query, xml_sheets, excel_data, excel_headers, max_results=12):
    q = (query or "").lower().strip()
    if not q: return {"xml_hits": [], "excel_hits": []}

    xml_hits, excel_hits = [], []

    primary = next(((n, v) for n, v in xml_sheets.items() if v["data"]), None)
    if primary:
        sheet_name, sheet = primary
        for idx, rec in enumerate(sheet["data"]):
            mf = [f for f in sheet["fields"]
                  if (rec.get(f) or "").lower() and
                  (q in (rec.get(f) or "").lower() or (rec.get(f) or "").lower() in q)]
            if mf: xml_hits.append({"idx": idx, "rec": rec, "matching_fields": mf})

    for idx, rec in enumerate(excel_data):
        mf = [h for h in excel_headers
              if (rec.get(h) or "").lower() and
              (q in (rec.get(h) or "").lower() or (rec.get(h) or "").lower() in q)]
        if mf: excel_hits.append({"idx": idx, "rec": rec, "matching_fields": mf})

    def score(hit):
        s = len(hit["matching_fields"])
        for f in hit["matching_fields"]:
            v = (hit["rec"].get(f) or "").lower()
            if v == q: s += 10
            elif v.startswith(q): s += 4
        return s

    return {
        "xml_hits":   sorted(xml_hits,   key=score, reverse=True)[:max_results],
        "excel_hits": sorted(excel_hits, key=score, reverse=True)[:max_results],
    }

# ─── Excel Report Export ──────────────────────────────────────────────────────
def export_excel_report(results, path):
    wb = openpyxl.Workbook()

    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="4F46E5")
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

    def auto_width(ws, min_w=10, max_w=40):
        for col in ws.columns:
            width = min_w
            for cell in col:
                try: width = min(max_w, max(width, len(str(cell.value or "")) + 2))
                except: pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

    s = results["summary"]
    join_cols = results["excel_join_cols"]

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Summary"
    ws1.append(["Genpact SAP Migration Validation Report"])
    ws1["A1"].font = Font(bold=True, size=14, color="4F46E5")
    ws1.append(["Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws1.append([])
    summary_data = [
        ("Total Records",           s["total"]),
        ("Matched",                 s["matched"]),
        ("Mismatches (Field Diffs)",s["mismatch"]),
        ("Missing in XML",          s["missing_xml"]),
        ("XML Only (not in Excel)", s["xml_only"]),
        ("Match Rate",              f"{s['match_rate']}%"),
    ]
    for label, val in summary_data:
        ws1.append([label, val])
    ws1.append([])
    ws1.append(["JOIN KEYS USED"])
    ws1["A" + str(ws1.max_row)].font = Font(bold=True)
    for e in join_cols:
        ws1.append([f"Excel: {e['excel_col']}", f"XML: {e['xml_key']}"])
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: All Records ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Records")
    key_cols  = [f"Key: {e['excel_col']}" for e in join_cols]
    fp_cols   = []
    for p in results["field_pairs"]:
        fp_cols += [f"Excel: {p['excel_col']}", f"XML: {p['xml_field']}", "Match?"]
    headers2 = ["#", "Status", "Key Display"] + key_cols + fp_cols
    style_header(ws2, 1, headers2)

    for i, r in enumerate(results["records"], 1):
        key_vals = [r["excel_rec"].get(e["excel_col"], "") for e in join_cols]
        fp_vals  = []
        for p in results["field_pairs"]:
            chk = next((c for c in r["field_checks"]
                        if c["excel_col"] == p["excel_col"]), None)
            if chk:
                fp_vals += [chk["excel_val"], chk["xml_val"], "YES" if chk["match"] else "NO"]
            else:
                fp_vals += ["", r["xml_rec"].get(p["xml_field"], "") if r["xml_rec"] else "N/A", ""]
        row_data = [i, r["status"].upper(), r["key_display"]] + key_vals + fp_vals
        ws2.append(row_data)
        fill = pass_fill if r["status"]=="matched" else (fail_fill if r["status"]=="missing_xml" else warn_fill)
        for c in range(1, 4): ws2.cell(ws2.max_row, c).fill = fill
    auto_width(ws2)

    # ── Sheet 3: Mismatches ───────────────────────────────────────────────────
    mismatch_rows = [(r["key_display"], c["excel_col"], c["xml_field"], c["excel_val"], c["xml_val"])
                     for r in results["records"] for c in r["field_checks"] if not c["match"]]
    if mismatch_rows:
        ws3 = wb.create_sheet("Mismatches")
        style_header(ws3, 1, ["Key", "Excel Column", "XML Field", "Excel Value", "XML Value"])
        for row in mismatch_rows: ws3.append(list(row))
        auto_width(ws3)

    # ── Sheet 4: XML Only ─────────────────────────────────────────────────────
    if results["xml_only"]:
        ws4 = wb.create_sheet("XML Only")
        fields = results["xml_fields"][:20]
        style_header(ws4, 1, fields)
        for rec in results["xml_only"]:
            ws4.append([rec.get(f, "") for f in fields])
        auto_width(ws4)

    wb.save(path)

# ─── AI Column Mapping via Claude ─────────────────────────────────────────────
def ai_match_columns(excel_headers, xml_fields, api_key):
    if not ANTHROPIC_OK or not api_key: return {}
    client = anthropic.Anthropic(api_key=api_key)
    prompt = (
        f"You are an SAP data migration expert. Map these Excel headers to SAP XML field codes.\n"
        f"Excel headers: {json.dumps(excel_headers)}\n"
        f"SAP XML fields: {json.dumps(xml_fields)}\n"
        f"Return ONLY a JSON object: {{\"Excel Header\": \"SAP_FIELD\" or null}}\n"
        f"Key mappings: Work Center→ARBPL, Plant→WERKS, Description→KTEXT, etc."
    )
    msg = client.messages.create(
        model="claude-sonnet-4-20250514", max_tokens=1000,
        messages=[{"role": "user", "content": prompt}]
    )
    text = msg.content[0].text
    try:
        clean = re.sub(r"```json|```", "", text).strip()
        return json.loads(clean)
    except: return {}

# ═══════════════════════════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════════════════════════
class SAPValidator(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Genpact SAP Validator — Post-Load Migration Validation")
        self.geometry("1280x800")
        self.minsize(900, 600)
        self.configure(bg=C["bg"])

        # State
        self.xml_sheets    = {}
        self.xml_path      = ""
        self.excel_data    = []
        self.excel_headers = []
        self.excel_path    = ""
        self.field_map     = {}
        self.join_keys     = []
        self.results       = None
        self.learned_maps  = []
        self.api_key       = os.environ.get("ANTHROPIC_API_KEY", "")

        self._build_ui()

    # ── UI construction ───────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg=C["border"],
                               sashwidth=4, sashrelief=tk.FLAT)
        paned.pack(fill=tk.BOTH, expand=True)
        self._build_sidebar(paned)
        self._build_main(paned)

    def _build_header(self):
        hdr = tk.Frame(self, bg=C["surface"], height=54)
        hdr.pack(fill=tk.X, side=tk.TOP)
        hdr.pack_propagate(False)
        tk.Frame(hdr, bg=C["border"], height=1).pack(fill=tk.X, side=tk.BOTTOM)

        left = tk.Frame(hdr, bg=C["surface"])
        left.pack(side=tk.LEFT, padx=16, pady=8)

        icon = tk.Label(left, text="⚙", bg=C["accent"], fg="white",
                        font=("Segoe UI", 14, "bold"), width=2, relief=tk.FLAT)
        icon.pack(side=tk.LEFT, padx=(0, 10))

        titf = tk.Frame(left, bg=C["surface"])
        titf.pack(side=tk.LEFT)
        tk.Label(titf, text="Genpact SAP Validator", bg=C["surface"],
                 fg=C["text"], font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(titf, text="Post-Load Migration Validation", bg=C["surface"],
                 fg=C["muted"], font=("Segoe UI", 9)).pack(anchor="w")

        right = tk.Frame(hdr, bg=C["surface"])
        right.pack(side=tk.RIGHT, padx=12)

        self.match_rate_lbl = tk.Label(right, text="", bg=C["warn_bg"],
                                       fg=C["warn"], font=("Segoe UI", 9, "bold"),
                                       padx=10, pady=3, relief=tk.FLAT)

        self.learned_lbl = tk.Label(right, text="", bg=C["teal_bg"],
                                    fg=C["teal"], font=("Segoe UI", 9, "bold"),
                                    padx=10, pady=3, relief=tk.FLAT)

        btn_style = dict(relief=tk.FLAT, font=("Segoe UI", 10, "bold"),
                         cursor="hand2", padx=12, pady=5)

        self.run_btn = tk.Button(right, text="▶  Run Validation",
                                 bg=C["accent"], fg="white",
                                 command=self._run_validation, **btn_style)
        self.run_btn.pack(side=tk.RIGHT, padx=(4, 0))

        tk.Button(right, text="↓  Export Excel",
                  bg=C["surface2"], fg=C["text"],
                  command=self._export_excel, **btn_style).pack(side=tk.RIGHT, padx=4)

        tk.Button(right, text="🔑  API Key",
                  bg=C["surface2"], fg=C["text"],
                  command=self._set_api_key, **btn_style).pack(side=tk.RIGHT, padx=4)

    def _build_sidebar(self, paned):
        sb = tk.Frame(paned, bg=C["surface"], width=270)
        sb.pack_propagate(False)
        paned.add(sb, minsize=220)

        tk.Label(sb, text="FILES & CONFIGURATION", bg=C["surface"], fg=C["muted"],
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=14, pady=(12, 4))

        canvas = tk.Canvas(sb, bg=C["surface"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(sb, orient="vertical", command=canvas.yview)
        self.sb_frame = tk.Frame(canvas, bg=C["surface"])
        self.sb_frame.bind("<Configure>",
                           lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.sb_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self._build_file_cards()

    def _build_file_cards(self):
        f = self.sb_frame

        # XML Card
        xml_card = self._card(f, "Step 1 — XML Template")
        tk.Button(xml_card, text="Browse XML file…",
                  bg=C["accent"], fg="white", relief=tk.FLAT,
                  font=("Segoe UI", 10, "bold"), cursor="hand2",
                  command=self._load_xml).pack(fill=tk.X, pady=(0,4))
        self.xml_lbl = tk.Label(xml_card, text="No file selected",
                                bg=C["surface2"], fg=C["muted"],
                                font=("Segoe UI", 9), wraplength=200, justify=tk.LEFT)
        self.xml_lbl.pack(fill=tk.X)
        self.xml_sheets_frame = tk.Frame(xml_card, bg=C["surface2"])
        self.xml_sheets_frame.pack(fill=tk.X, pady=(4, 0))

        # Excel Card
        xl_card = self._card(f, "Step 2 — Post-Load Excel (.xlsx)")
        tk.Button(xl_card, text="Browse Excel file…",
                  bg=C["accent"], fg="white", relief=tk.FLAT,
                  font=("Segoe UI", 10, "bold"), cursor="hand2",
                  command=self._load_excel).pack(fill=tk.X, pady=(0,4))
        self.xl_lbl = tk.Label(xl_card, text="No file selected",
                               bg=C["surface2"], fg=C["muted"],
                               font=("Segoe UI", 9), wraplength=200, justify=tk.LEFT)
        self.xl_lbl.pack(fill=tk.X)

        # Join Keys Card
        jk_card = self._card(f, "Join Keys")
        self.jk_frame = tk.Frame(jk_card, bg=C["surface2"])
        self.jk_frame.pack(fill=tk.X, pady=(0, 4))
        tk.Button(jk_card, text="✏ Edit Join Keys",
                  bg=C["accent_l"], fg=C["accent"], relief=tk.FLAT,
                  font=("Segoe UI", 9, "bold"), cursor="hand2",
                  command=self._edit_join_keys).pack(fill=tk.X)

        # Field Map Card
        fm_card = self._card(f, "Field Mapping")
        btn_row = tk.Frame(fm_card, bg=C["surface2"])
        btn_row.pack(fill=tk.X, pady=(0, 4))
        tk.Button(btn_row, text="✏ Edit Mapping",
                  bg=C["accent_l"], fg=C["accent"], relief=tk.FLAT,
                  font=("Segoe UI", 9, "bold"), cursor="hand2",
                  command=self._edit_field_map).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0,2))
        tk.Button(btn_row, text="🤖 AI Map",
                  bg="#7c3aed20", fg="#7c3aed", relief=tk.FLAT,
                  font=("Segoe UI", 9, "bold"), cursor="hand2",
                  command=self._ai_map).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2,0))
        self.fm_frame = tk.Frame(fm_card, bg=C["surface2"])
        self.fm_frame.pack(fill=tk.X)

    def _card(self, parent, title):
        outer = tk.Frame(parent, bg=C["border"], pady=1, padx=1)
        outer.pack(fill=tk.X, padx=8, pady=4)
        inner = tk.Frame(outer, bg=C["surface2"], padx=10, pady=8)
        inner.pack(fill=tk.X)
        tk.Label(inner, text=title.upper(), bg=C["surface2"], fg=C["muted"],
                 font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(0,6))
        return inner

    def _build_main(self, paned):
        main = tk.Frame(paned, bg=C["bg"])
        paned.add(main, minsize=600)

        # Status bar
        self.status_bar = tk.Label(main, text="Upload XML and Excel files to begin.",
                                   bg=C["info_bg"], fg=C["info"],
                                   font=("Segoe UI", 10), anchor="w", padx=12, pady=6)
        self.status_bar.pack(fill=tk.X, padx=12, pady=(12, 0))

        # Summary cards row
        self.cards_frame = tk.Frame(main, bg=C["bg"])
        self.cards_frame.pack(fill=tk.X, padx=12, pady=8)

        # Tabs
        tab_bar = tk.Frame(main, bg=C["bg"])
        tab_bar.pack(fill=tk.X, padx=12)
        self.tab_btns = {}
        for key, label in [("all","All Records"),("mismatch","Mismatches"),
                           ("missing","Missing in XML"),("xmlonly","XML Only")]:
            b = tk.Button(tab_bar, text=label, relief=tk.FLAT, cursor="hand2",
                          font=("Segoe UI", 9, "bold"), padx=14, pady=6,
                          command=lambda k=key: self._switch_tab(k))
            b.pack(side=tk.LEFT)
            self.tab_btns[key] = b
        self._switch_tab("all")

        # Search & filter
        fbar = tk.Frame(main, bg=C["bg"])
        fbar.pack(fill=tk.X, padx=12, pady=(4, 0))

        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *a: self._refresh_table())
        tk.Label(fbar, text="Search:", bg=C["bg"], fg=C["muted"],
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Entry(fbar, textvariable=self.search_var, width=28,
                 font=("Segoe UI", 10), relief=tk.FLAT,
                 bg=C["surface"], fg=C["text"]).pack(side=tk.LEFT, padx=6)

        self.filter_var = tk.StringVar(value="all")
        for val, lbl in [("all","All"),("matched","Matched"),
                         ("mismatch","Diffs"),("missing_xml","Missing")]:
            tk.Radiobutton(fbar, text=lbl, variable=self.filter_var, value=val,
                           bg=C["bg"], fg=C["text"], font=("Segoe UI", 9),
                           selectcolor=C["accent_l"],
                           command=self._refresh_table).pack(side=tk.LEFT, padx=4)

        self.count_lbl = tk.Label(fbar, text="", bg=C["bg"], fg=C["muted"],
                                  font=("Segoe UI", 9))
        self.count_lbl.pack(side=tk.RIGHT)

        # Results treeview
        tree_frame = tk.Frame(main, bg=C["bg"])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        self.tree = ttk.Treeview(tree_frame, show="headings",
                                 selectmode="browse", style="SAP.Treeview")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<Double-1>", self._on_row_double_click)
        self.tree.bind("<Return>", self._on_row_double_click)

        # Detail pane
        self.detail_frame = tk.Frame(main, bg=C["surface"],
                                     relief=tk.FLAT, bd=1)
        self.detail_text = tk.Text(self.detail_frame, height=8, state=tk.DISABLED,
                                   font=("Consolas", 9), bg=C["surface2"],
                                   fg=C["text"], relief=tk.FLAT, wrap=tk.WORD,
                                   padx=10, pady=8)
        self.detail_text.pack(fill=tk.BOTH, expand=True)
        self.detail_frame.pack(fill=tk.X, padx=12, pady=(0,8))

        self._style_treeview()

    def _style_treeview(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("SAP.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], rowheight=26,
                        font=("Segoe UI", 10))
        style.configure("SAP.Treeview.Heading",
                        background=C["surface2"], foreground=C["muted"],
                        font=("Segoe UI", 9, "bold"), relief=tk.FLAT)
        style.map("SAP.Treeview",
                  background=[("selected", C["accent_l"])],
                  foreground=[("selected", C["accent"])])
        self.tree.tag_configure("matched",     background=C["pass_bg"],  foreground=C["pass_"])
        self.tree.tag_configure("mismatch",    background=C["warn_bg"],  foreground=C["warn"])
        self.tree.tag_configure("missing_xml", background=C["fail_bg"],  foreground=C["fail"])
        self.tree.tag_configure("xml_only",    background=C["info_bg"],  foreground=C["info"])
        self.tree.tag_configure("odd",         background=C["surface2"])
        self.tree.tag_configure("even",        background=C["surface"])

    # ── File loading ──────────────────────────────────────────────────────────
    def _load_xml(self):
        path = filedialog.askopenfilename(title="Select XML Migration Template",
                                         filetypes=[("XML files","*.xml"),("All","*.*")])
        if not path: return
        try:
            self.xml_sheets = parse_xml_workbook(path)
            self.xml_path   = path
            name = os.path.basename(path)
            self.xml_lbl.config(text=f"✓ {name}", fg=C["pass_"])
            self._refresh_xml_sheets_display()
            self._auto_map()
            self._status(f"XML loaded: {len(self.xml_sheets)} sheets", "info")
        except Exception as e:
            messagebox.showerror("XML Error", str(e))

    def _load_excel(self):
        path = filedialog.askopenfilename(title="Select Post-Load Excel",
                                         filetypes=[("Excel files","*.xlsx *.xls"),("All","*.*")])
        if not path: return
        try:
            parsed = parse_excel_file(path)
            self.excel_data    = parsed["data"]
            self.excel_headers = parsed["headers"]
            self.excel_path    = path
            name = os.path.basename(path)
            self.xl_lbl.config(
                text=f"✓ {name}\n{len(self.excel_data):,} rows · {len(self.excel_headers)} columns",
                fg=C["pass_"])
            self._auto_map()
            self._status(f"Excel loaded: {len(self.excel_data):,} rows", "info")
        except Exception as e:
            messagebox.showerror("Excel Error", str(e))

    def _auto_map(self):
        if not self.xml_sheets or not self.excel_headers: return
        primary = next((v for v in self.xml_sheets.values() if v["data"]), None)
        if not primary: return
        self.field_map = build_field_map(self.excel_headers, primary["fields"])
        self.join_keys = detect_join_keys(primary["fields"])
        self._refresh_jk_display()
        self._refresh_fm_display()

    def _refresh_xml_sheets_display(self):
        for w in self.xml_sheets_frame.winfo_children(): w.destroy()
        for name, sheet in self.xml_sheets.items():
            color = C["teal"] if sheet["data"] else C["muted"]
            bg    = C["teal_bg"] if sheet["data"] else C["surface2"]
            tk.Label(self.xml_sheets_frame, text=f"{name} ({len(sheet['data'])})",
                     bg=bg, fg=color, font=("Segoe UI", 8, "bold"),
                     padx=6, pady=2, relief=tk.FLAT).pack(side=tk.LEFT, padx=2, pady=2)

    def _refresh_jk_display(self):
        for w in self.jk_frame.winfo_children(): w.destroy()
        if not self.join_keys:
            tk.Label(self.jk_frame, text="No join keys set", bg=C["surface2"],
                     fg=C["warn"], font=("Segoe UI", 9)).pack()
            return
        for i, k in enumerate(self.join_keys):
            if i > 0:
                tk.Label(self.jk_frame, text="+", bg=C["surface2"],
                         fg=C["muted"], font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
            tk.Label(self.jk_frame, text=k, bg=C["accent_l"], fg=C["accent"],
                     font=("Segoe UI", 9, "bold"), padx=8, pady=3,
                     relief=tk.FLAT).pack(side=tk.LEFT, padx=2)

    def _refresh_fm_display(self):
        for w in self.fm_frame.winfo_children(): w.destroy()
        mapped   = sum(1 for v in self.field_map.values() if v)
        total    = len(self.field_map)
        tk.Label(self.fm_frame,
                 text=f"{mapped}/{total} fields mapped",
                 bg=C["surface2"], fg=C["accent"],
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0,4))
        for h, sap in list(self.field_map.items())[:12]:
            row = tk.Frame(self.fm_frame, bg=C["surface2"])
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=h[:22], bg=C["surface2"], fg=C["text"],
                     font=("Segoe UI", 8), width=22, anchor="w").pack(side=tk.LEFT)
            tk.Label(row, text=sap or "—", bg=C["surface2"],
                     fg=C["accent"] if sap else C["muted"],
                     font=("Segoe UI", 8, "bold")).pack(side=tk.RIGHT)
        if total > 12:
            tk.Label(self.fm_frame, text=f"+ {total-12} more…",
                     bg=C["surface2"], fg=C["muted"], font=("Segoe UI", 8)).pack(anchor="w")

    # ── Validation ────────────────────────────────────────────────────────────
    def _run_validation(self):
        if not self.xml_sheets:
            messagebox.showwarning("Missing File", "Please upload the XML template first."); return
        if not self.excel_data:
            messagebox.showwarning("Missing File", "Please upload the post-load Excel first."); return
        if not self.join_keys:
            messagebox.showwarning("No Join Keys", "Please set at least one join key."); return

        self._status("Running validation…", "info")
        self.run_btn.config(state=tk.DISABLED)

        def _do():
            r = run_validation(self.xml_sheets, self.excel_data, self.excel_headers,
                               self.join_keys, self.field_map)
            self.after(0, lambda: self._show_results(r))

        threading.Thread(target=_do, daemon=True).start()

    def _show_results(self, results):
        self.results = results
        self.run_btn.config(state=tk.NORMAL)
        s = results["summary"]
        rate = s["match_rate"]
        col  = C["pass_"] if rate==100 else C["warn"] if rate>=80 else C["fail"]
        bg   = C["pass_bg"] if rate==100 else C["warn_bg"] if rate>=80 else C["fail_bg"]

        self._status(
            f"Validation complete — {rate}% match rate  |  "
            f"{s['matched']:,} matched  |  {s['mismatch']:,} diffs  |  "
            f"{s['missing_xml']:,} missing in XML  |  {s['xml_only']:,} XML only",
            "pass" if rate==100 else "warn"
        )
        self.match_rate_lbl.config(
            text=f"Match Rate: {rate}%", bg=bg, fg=col)
        self.match_rate_lbl.pack(side=tk.RIGHT, padx=4)
        self._build_summary_cards(s)
        self._update_tab_labels(s)
        self._refresh_table()

    def _build_summary_cards(self, s):
        for w in self.cards_frame.winfo_children(): w.destroy()
        cards = [
            ("Total", s["total"],        C["text"],  C["surface"]),
            ("Matched", s["matched"],     C["pass_"], C["pass_bg"]),
            ("Diffs",   s["mismatch"],    C["warn"],  C["warn_bg"]),
            ("Missing", s["missing_xml"], C["fail"],  C["fail_bg"]),
            ("XML Only",s["xml_only"],    C["info"],  C["info_bg"]),
            ("Match %", f"{s['match_rate']}%", C["accent"], C["accent_l"]),
        ]
        for label, val, fg, bg in cards:
            card = tk.Frame(self.cards_frame, bg=bg, relief=tk.FLAT, bd=1)
            card.pack(side=tk.LEFT, padx=4, fill=tk.Y)
            tk.Label(card, text=str(val), bg=bg, fg=fg,
                     font=("Segoe UI", 18, "bold"), padx=16, pady=8).pack()
            tk.Label(card, text=label, bg=bg, fg=fg,
                     font=("Segoe UI", 8), padx=16, pady=(0,8)).pack()

    def _update_tab_labels(self, s):
        self.tab_btns["all"].config(text=f"All Records ({s['total']})")
        self.tab_btns["mismatch"].config(text=f"Mismatches ({s['mismatch']})")
        self.tab_btns["missing"].config(text=f"⚠ Missing in XML ({s['missing_xml']})")
        self.tab_btns["xmlonly"].config(text=f"XML Only ({s['xml_only']})")

    def _switch_tab(self, key):
        self.active_tab = key
        for k, b in self.tab_btns.items():
            b.config(bg=C["accent"] if k==key else C["surface2"],
                     fg="white" if k==key else C["muted"])
        self._refresh_table()

    # ── Table rendering ───────────────────────────────────────────────────────
    def _refresh_table(self):
        self.tree.delete(*self.tree.get_children())
        if not self.results: return

        tab    = getattr(self, "active_tab", "all")
        sq     = self.search_var.get().lower()
        filt   = self.filter_var.get()
        r      = self.results

        if tab == "xmlonly":
            self._render_xml_only(r["xml_only"])
            return

        recs = r["records"]
        if tab == "mismatch":    recs = [x for x in recs if x["status"]=="mismatch"]
        elif tab == "missing":   recs = [x for x in recs if x["status"]=="missing_xml"]
        elif filt != "all":      recs = [x for x in recs if x["status"]==filt]
        if sq: recs = [x for x in recs if sq in x["key_display"].lower()]

        # Build columns dynamically
        join_cols = [e["excel_col"] for e in r["excel_join_cols"]]
        fp_cols   = [p["excel_col"] for p in r["field_pairs"]]
        all_cols  = ["#","Status","Key"] + join_cols + fp_cols[:8]

        self.tree["columns"] = all_cols
        for col in all_cols:
            w = 50 if col=="#" else 100 if col=="Status" else 160 if col=="Key" else 110
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=40, anchor="w")

        for i, rec in enumerate(recs):
            key_vals = [rec["excel_rec"].get(c,"") for c in join_cols]
            fp_vals  = []
            for p in r["field_pairs"][:8]:
                chk = next((c for c in rec["field_checks"] if c["excel_col"]==p["excel_col"]), None)
                if chk:
                    fp_vals.append(("✓ " if chk["match"] else "✗ ") + chk["excel_val"][:18])
                else:
                    fp_vals.append("")
            values = [i+1, rec["status"].upper(), rec["key_display"]] + key_vals + fp_vals
            tags = (rec["status"], "odd" if i%2 else "even")
            self.tree.insert("", tk.END, iid=str(i), values=values, tags=tags)

        self.count_lbl.config(text=f"{len(recs)} records")

    def _render_xml_only(self, xml_only):
        if not self.results: return
        fields = self.results["xml_fields"][:10]
        cols   = ["#"] + fields
        self.tree["columns"] = cols
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=50 if col=="#" else 120, anchor="w")
        for i, rec in enumerate(xml_only):
            vals = [i+1] + [rec.get(f,"") for f in fields]
            self.tree.insert("", tk.END, iid=str(i), values=vals,
                             tags=("xml_only", "odd" if i%2 else "even"))
        self.count_lbl.config(text=f"{len(xml_only)} records")

    def _on_row_double_click(self, event):
        sel = self.tree.selection()
        if not sel or not self.results: return
        idx = int(sel[0])
        tab = getattr(self, "active_tab", "all")

        if tab == "xmlonly":
            rec = self.results["xml_only"][idx]
            self._show_detail_text(f"XML ONLY RECORD\n{'='*60}\n" +
                                   "\n".join(f"  {k}: {v}" for k,v in rec.items() if v))
            return

        recs = self.results["records"]
        if tab == "mismatch":  recs = [r for r in recs if r["status"]=="mismatch"]
        elif tab == "missing": recs = [r for r in recs if r["status"]=="missing_xml"]

        if idx >= len(recs): return
        rec = recs[idx]

        if rec["status"] == "missing_xml":
            self._open_smart_search(rec)
        else:
            lines = [f"RECORD: {rec['key_display']}",
                     f"STATUS: {rec['status'].upper()}",
                     "="*60]
            for chk in rec["field_checks"]:
                mark = "✓" if chk["match"] else "✗"
                lines.append(f"  {mark} {chk['excel_col']} ({chk['xml_field']})")
                if not chk["match"]:
                    lines.append(f"      Excel: {chk['excel_val']!r}")
                    lines.append(f"      XML:   {chk['xml_val']!r}")
            self._show_detail_text("\n".join(lines))

    def _show_detail_text(self, text):
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.delete("1.0", tk.END)
        self.detail_text.insert(tk.END, text)
        self.detail_text.config(state=tk.DISABLED)

    # ── Smart Search & Learn ──────────────────────────────────────────────────
    def _open_smart_search(self, record):
        win = tk.Toplevel(self)
        win.title("🔍 Smart Key Search & Learn")
        win.geometry("820x640")
        win.configure(bg=C["bg"])
        win.transient(self)
        win.grab_set()

        # Header
        hdr = tk.Frame(win, bg=C["surface"], pady=12, padx=16)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🔍  Smart Key Search & Learn",
                 bg=C["surface"], fg=C["text"],
                 font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(hdr, text=f"Missing key: {record['key_display']}",
                 bg=C["surface"], fg=C["fail"],
                 font=("Segoe UI", 10)).pack(anchor="w")
        tk.Frame(win, bg=C["border"], height=1).pack(fill=tk.X)

        body = tk.Frame(win, bg=C["bg"])
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        # Info banner
        info = tk.Label(body,
            text="Type any value from this record to search both files. "
                 "When you find the XML match, click 'This is the match' — "
                 "the tool learns the correct join field automatically.",
            bg=C["accent_l"], fg=C["accent"],
            font=("Segoe UI", 9), wraplength=760, justify=tk.LEFT,
            padx=12, pady=8, relief=tk.FLAT)
        info.pack(fill=tk.X, pady=(0, 10))

        # Excel record display
        tk.Label(body, text="EXCEL RECORD (NOT FOUND IN XML)",
                 bg=C["bg"], fg=C["muted"],
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        rec_frame = tk.Frame(body, bg=C["fail_bg"], relief=tk.FLAT, pady=8, padx=10)
        rec_frame.pack(fill=tk.X, pady=(2,8))
        chip_row = tk.Frame(rec_frame, bg=C["fail_bg"])
        chip_row.pack(fill=tk.X)
        for k, v in list(record["excel_rec"].items())[:12]:
            if not v: continue
            chip = tk.Frame(chip_row, bg=C["surface"], relief=tk.FLAT, padx=6, pady=3)
            chip.pack(side=tk.LEFT, padx=3, pady=2)
            tk.Label(chip, text=k, bg=C["surface"], fg=C["muted"],
                     font=("Segoe UI", 7)).pack(anchor="w")
            tk.Label(chip, text=str(v)[:25], bg=C["surface"], fg=C["text"],
                     font=("Segoe UI", 9, "bold")).pack(anchor="w")

        # Search bar
        tk.Label(body, text="SEARCH IN BOTH FILES",
                 bg=C["bg"], fg=C["muted"],
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        search_var = tk.StringVar()
        entry = tk.Entry(body, textvariable=search_var, font=("Segoe UI", 11),
                         bg=C["surface"], fg=C["text"], relief=tk.FLAT,
                         insertbackground=C["text"])
        entry.pack(fill=tk.X, pady=(2, 4), ipady=5)
        entry.focus_set()

        # Quick-fill chips
        qrow = tk.Frame(body, bg=C["bg"])
        qrow.pack(fill=tk.X, pady=(0,8))
        tk.Label(qrow, text="Quick search:", bg=C["bg"], fg=C["muted"],
                 font=("Segoe UI", 8)).pack(side=tk.LEFT)
        for k, v in list(record["excel_rec"].items())[:8]:
            if not v or len(str(v)) > 30: continue
            def _fill(val=v): search_var.set(val); _do_search()
            tk.Button(qrow, text=str(v), bg=C["surface2"], fg=C["text"],
                      font=("Segoe UI", 8), relief=tk.FLAT, cursor="hand2",
                      command=_fill).pack(side=tk.LEFT, padx=2)

        # Results area
        results_frame = tk.Frame(body, bg=C["bg"])
        results_frame.pack(fill=tk.BOTH, expand=True)

        xml_lbl   = tk.Label(results_frame, text="XML matches: 0",
                             bg=C["bg"], fg=C["muted"], font=("Segoe UI", 8,"bold"))
        xml_lbl.pack(anchor="w")
        xml_tree = ttk.Treeview(results_frame, show="headings", height=5,
                                style="SAP.Treeview", selectmode="browse")
        xml_tree.pack(fill=tk.X, pady=(2,6))

        excel_lbl = tk.Label(results_frame, text="Excel matches: 0",
                             bg=C["bg"], fg=C["muted"], font=("Segoe UI", 8,"bold"))
        excel_lbl.pack(anchor="w")
        excel_tree = ttk.Treeview(results_frame, show="headings", height=4,
                                  style="SAP.Treeview", selectmode="browse")
        excel_tree.pack(fill=tk.X, pady=(2,6))

        # Status / learned notice
        notice_lbl = tk.Label(body, text="", bg=C["bg"], fg=C["pass_"],
                              font=("Segoe UI", 9, "bold"), wraplength=760)
        notice_lbl.pack(anchor="w")

        def _do_search(*_):
            q = search_var.get().strip()
            hits = fuzzy_search(q, self.xml_sheets, self.excel_data,
                                self.excel_headers)

            # Populate XML tree
            xml_tree.delete(*xml_tree.get_children())
            if hits["xml_hits"]:
                primary = next((v for v in self.xml_sheets.values() if v["data"]), None)
                fields  = (primary["fields"][:8] if primary else []) or []
                xml_tree["columns"] = ["Row","Matched On"] + fields
                xml_tree.heading("Row", text="#")
                xml_tree.column("Row", width=40)
                xml_tree.heading("Matched On", text="Matched Field")
                xml_tree.column("Matched On", width=120)
                for f in fields:
                    xml_tree.heading(f, text=f)
                    xml_tree.column(f, width=100)
                for hit in hits["xml_hits"]:
                    vals = ([hit["idx"]+1,
                             ", ".join(hit["matching_fields"])] +
                            [hit["rec"].get(f,"") for f in fields])
                    xml_tree.insert("", tk.END, values=vals)
            xml_lbl.config(text=f"XML matches: {len(hits['xml_hits'])}")

            # Populate Excel tree
            excel_tree.delete(*excel_tree.get_children())
            if hits["excel_hits"]:
                cols = ["Row","Matched On"] + self.excel_headers[:6]
                excel_tree["columns"] = cols
                excel_tree.heading("Row", text="#"); excel_tree.column("Row",width=40)
                excel_tree.heading("Matched On",text="Matched"); excel_tree.column("Matched On",width=120)
                for h in self.excel_headers[:6]:
                    excel_tree.heading(h,text=h); excel_tree.column(h,width=90)
                for hit in hits["excel_hits"]:
                    vals = [hit["idx"]+1, ", ".join(hit["matching_fields"])] + \
                           [hit["rec"].get(h,"") for h in self.excel_headers[:6]]
                    excel_tree.insert("",tk.END, values=vals)
            excel_lbl.config(text=f"Excel matches: {len(hits['excel_hits'])}")

        search_var.trace("w", _do_search)

        def _apply_match():
            sel = xml_tree.selection()
            if not sel:
                messagebox.showinfo("Select a match",
                    "Click on an XML row that corresponds to this Excel record first.")
                return
            item = xml_tree.item(sel[0])
            row_idx = int(item["values"][0]) - 1
            primary = next((v for v in self.xml_sheets.values() if v["data"]), None)
            if not primary: return
            xml_rec = primary["data"][row_idx]
            mf_str  = str(item["values"][1])
            xml_match_field = mf_str.split(",")[0].strip()

            # Find which Excel column has the same value as this XML field
            xml_val = (xml_rec.get(xml_match_field) or "").lower().strip()
            excel_match_field = next(
                (h for h in self.excel_headers
                 if (record["excel_rec"].get(h) or "").lower().strip() == xml_val),
                None
            ) or self.excel_headers[0]

            # Learn: update field map and join keys
            self.field_map[excel_match_field] = xml_match_field
            if xml_match_field not in self.join_keys:
                self.join_keys.append(xml_match_field)

            label = f"{excel_match_field} → {xml_match_field}"
            if label not in [m["label"] for m in self.learned_maps]:
                self.learned_maps.append({"excel_field": excel_match_field,
                                          "xml_field": xml_match_field,
                                          "label": label})

            # Update header badge
            self.learned_lbl.config(text=f"🧠 {len(self.learned_maps)} learned mapping(s)")
            self.learned_lbl.pack(side=tk.RIGHT, padx=4)

            self._refresh_jk_display()
            self._refresh_fm_display()

            notice_lbl.config(
                text=f"✓ Learned: '{excel_match_field}' (Excel) → '{xml_match_field}' (XML). "
                     f"Close this window and click Run Validation to apply.",
                bg=C["pass_bg"])

        tk.Button(body, text="✓  This is the match — learn this key",
                  bg=C["pass_"], fg="white", font=("Segoe UI", 10, "bold"),
                  relief=tk.FLAT, cursor="hand2", padx=14, pady=6,
                  command=_apply_match).pack(anchor="w", pady=(4,0))

    # ── Dialogs ───────────────────────────────────────────────────────────────
    def _edit_join_keys(self):
        primary = next((v for v in self.xml_sheets.values() if v["data"]), None)
        if not primary:
            messagebox.showinfo("Load XML first", "Upload the XML template first."); return

        win = tk.Toplevel(self)
        win.title("Edit Join Keys")
        win.geometry("500x500")
        win.configure(bg=C["bg"])
        win.transient(self); win.grab_set()

        tk.Label(win, text="Edit Join Keys", bg=C["bg"], fg=C["text"],
                 font=("Segoe UI", 12, "bold")).pack(pady=(14,4))
        tk.Label(win, text="Select which XML fields to use as the composite join key:",
                 bg=C["bg"], fg=C["muted"], font=("Segoe UI", 9)).pack()

        # Listbox with checkboxes (use Listbox + vars)
        frame = tk.Frame(win, bg=C["surface"], relief=tk.FLAT, bd=1)
        frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)
        lb_frame = tk.Frame(frame, bg=C["surface"])
        lb_frame.pack(fill=tk.BOTH, expand=True)

        vars_ = {}
        for f in primary["fields"]:
            var = tk.BooleanVar(value=f in self.join_keys)
            vars_[f] = var
            row = tk.Frame(lb_frame, bg=C["surface"])
            row.pack(fill=tk.X, padx=8, pady=1)
            tk.Checkbutton(row, text=f"{f}  ({SAP_LABELS.get(f,'')})",
                           variable=var, bg=C["surface"], fg=C["text"],
                           font=("Segoe UI", 10), selectcolor=C["accent_l"],
                           activebackground=C["surface"]).pack(anchor="w")

        def _save():
            new_keys = [f for f in primary["fields"] if vars_.get(f, tk.BooleanVar()).get()]
            if not new_keys:
                messagebox.showwarning("No keys", "Select at least one join key."); return
            self.join_keys = new_keys
            self._refresh_jk_display()
            win.destroy()

        tk.Button(win, text="Apply", bg=C["accent"], fg="white",
                  font=("Segoe UI", 10, "bold"), relief=tk.FLAT, cursor="hand2",
                  padx=16, pady=6, command=_save).pack(pady=8)

    def _edit_field_map(self):
        if not self.excel_headers:
            messagebox.showinfo("Load Excel first", "Upload the post-load Excel first."); return
        primary = next((v for v in self.xml_sheets.values() if v["data"]), None)
        if not primary:
            messagebox.showinfo("Load XML first", "Upload the XML template first."); return

        win = tk.Toplevel(self)
        win.title("Edit Field Mapping")
        win.geometry("600x560")
        win.configure(bg=C["bg"])
        win.transient(self); win.grab_set()

        tk.Label(win, text="Edit Field Mapping", bg=C["bg"], fg=C["text"],
                 font=("Segoe UI", 12, "bold")).pack(pady=(14,2))
        tk.Label(win, text="Map each Excel column → SAP XML field code (or leave as — to skip)",
                 bg=C["bg"], fg=C["muted"], font=("Segoe UI", 9)).pack()

        canvas = tk.Canvas(win, bg=C["surface"], highlightthickness=0)
        sb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=C["surface"])
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=16, pady=8)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        xml_opts = [""] + primary["fields"]
        vars_ = {}
        for h in self.excel_headers:
            row = tk.Frame(inner, bg=C["surface"])
            row.pack(fill=tk.X, padx=8, pady=2)
            tk.Label(row, text=h[:30], bg=C["surface"], fg=C["text"],
                     font=("Segoe UI", 9), width=28, anchor="w").pack(side=tk.LEFT)
            var = tk.StringVar(value=self.field_map.get(h) or "")
            vars_[h] = var
            cb = ttk.Combobox(row, textvariable=var, values=xml_opts,
                              width=18, font=("Segoe UI", 9))
            cb.pack(side=tk.LEFT, padx=6)

        def _save():
            for h, var in vars_.items():
                v = var.get().strip()
                self.field_map[h] = v if v else None
            self._refresh_fm_display()
            win.destroy()

        tk.Button(win, text="Apply", bg=C["accent"], fg="white",
                  font=("Segoe UI", 10, "bold"), relief=tk.FLAT, cursor="hand2",
                  padx=16, pady=6, command=_save).pack(pady=6)

    def _ai_map(self):
        if not ANTHROPIC_OK:
            messagebox.showerror("Missing Package",
                "Install anthropic: pip install anthropic"); return
        if not self.api_key:
            self._set_api_key()
            if not self.api_key: return
        primary = next((v for v in self.xml_sheets.values() if v["data"]), None)
        if not primary: return

        self._status("Calling Claude AI for column mapping…", "info")
        def _do():
            new_map = ai_match_columns(self.excel_headers, primary["fields"], self.api_key)
            def _apply():
                for h, v in new_map.items():
                    if not self.field_map.get(h) and v:
                        self.field_map[h] = v
                self._refresh_fm_display()
                self._status("AI mapping applied!", "pass")
            self.after(0, _apply)
        threading.Thread(target=_do, daemon=True).start()

    def _set_api_key(self):
        key = simpledialog.askstring("Anthropic API Key",
            "Enter your Anthropic API key (for AI column mapping):\n"
            "Get one at console.anthropic.com",
            initialvalue=self.api_key, show="*", parent=self)
        if key: self.api_key = key.strip()

    # ── Export ────────────────────────────────────────────────────────────────
    def _export_excel(self):
        if not self.results:
            messagebox.showinfo("No results", "Run validation first."); return
        if not OPENPYXL_OK:
            messagebox.showerror("Missing Package", "pip install openpyxl"); return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"SAP_Validation_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not path: return
        try:
            export_excel_report(self.results, path)
            messagebox.showinfo("Export Complete",
                f"Report saved to:\n{path}\n\n4 sheets: Summary, All Records, Mismatches, XML Only")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ── Status bar ────────────────────────────────────────────────────────────
    def _status(self, msg, level="info"):
        colors = {
            "info": (C["info_bg"],  C["info"]),
            "pass": (C["pass_bg"], C["pass_"]),
            "warn": (C["warn_bg"], C["warn"]),
            "fail": (C["fail_bg"], C["fail"]),
        }
        bg, fg = colors.get(level, (C["info_bg"], C["info"]))
        self.status_bar.config(text=msg, bg=bg, fg=fg)


# ─── Entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not OPENPYXL_OK:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl lxml anthropic")
        exit(1)
    app = SAPValidator()
    app.mainloop()
